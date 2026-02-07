import os
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

# Create your views here.
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response

from .authentication import IsJWTAuthenticated


from django.contrib.auth import authenticate
from django.db.models import Count, Q, F, Max
from django.utils import timezone
from django.conf import settings
from django.core.cache import cache
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import random
import string
import hashlib


def generate_alphanumeric_otp(length=4):
    """Generate alphanumeric OTP (uppercase letters and digits, excluding confusing chars)"""
    # Exclude 0, O, I, 1 to avoid confusion
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(random.choice(chars) for _ in range(length))
from google.oauth2 import id_token
from google.auth.transport import requests
import jwt
from datetime import timedelta
import firebase_admin
from firebase_admin import credentials, auth as firebase_auth

from .models import (
    Department, Teacher, Student, Class as ClassModel, Session, Attendance, Device, Period, Subject,
    StudentClass, StudentInvitation, Course, Organization, OrganizationAdmin,
    Announcement, AnnouncementRead, TeacherNote, PeriodTimeSlot, ClassTeacher, SubjectEnrollment
)
from .serializers import (
    InitiateAttendanceSerializer, InitiateAttendanceResponseSerializer,
    SubmitOTPSerializer, SubmitOTPResponseSerializer, ManualMarkSerializer,
    LiveStatusSerializer, AttendanceSerializer, StudentSummarySerializer,
    ClassSummarySerializer, GoogleAuthSerializer, AuthResponseSerializer,
    AdminLoginSerializer, VerifyUserSerializer, ClassDetailSerializer,
    ClassCreateSerializer, StudentEnrollmentSerializer, StudentInviteSerializer,
    TeacherProfileSerializer, StudentProfileSerializer, PeriodSerializer,
    SubjectSerializer, DepartmentSerializer, PeriodCreateSerializer, SubjectCreateSerializer,
    CourseCreateSerializer, CourseSerializer
)
import csv
import io
import calendar
from collections import defaultdict

if not firebase_admin._apps:
    firebase_key_json = os.getenv('FIREBASE_KEY_JSON', '')
    if firebase_key_json:
        import json
        cred = credentials.Certificate(json.loads(firebase_key_json))
    else:
        firebase_key_path = os.getenv('FIREBASE_KEY_PATH', 'firebase-key.json')
        cred = credentials.Certificate(firebase_key_path)
    firebase_admin.initialize_app(cred)

def sanitize_group_name(email):
    """Convert email to valid group name"""
    return hashlib.md5(email.encode()).hexdigest()[:30]


def _encode_jwt(email, user_type, device_id=None, refresh=False):
    expiry = timedelta(days=7 if refresh else 1)
    payload = {
        'email': email,
        'user_type': user_type,
        'exp': timezone.now() + expiry,
        'iat': timezone.now(),
        'type': 'refresh' if refresh else 'access'
    }

    if user_type == 'student':
        payload['device_id'] = device_id

    return jwt.encode(payload, getattr(settings, 'JWT_SECRET', settings.SECRET_KEY), algorithm='HS256')


class AdminLoginView(APIView):
    permission_classes = []

    def post(self, request):
        serializer = AdminLoginSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user = authenticate(
            request,
            username=serializer.validated_data['username'],
            password=serializer.validated_data['password'],
        )

        if not user or not (user.is_staff or user.is_superuser):
            return Response({'error': 'Invalid admin credentials'}, status=status.HTTP_401_UNAUTHORIZED)

        access_token = _encode_jwt(user.username, 'admin')
        refresh_token = _encode_jwt(user.username, 'admin', refresh=True)

        return Response(
            {
                'access_token': access_token,
                'refresh_token': refresh_token,
                'user_type': 'admin',
                'user_info': {
                    'username': user.username,
                    'email': user.email,
                }
            },
            status=status.HTTP_200_OK,
        )


class PendingUsersView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'admin':
            return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)

        user_type = request.query_params.get('user_type')

        if user_type == 'teacher':
            teachers = Teacher.objects.filter(verified=False)
            data = [
                {
                    'user_type': 'teacher',
                    'email': t.teacher_email,
                    'name': t.name,
                    'designation': t.designation,
                    'department_id': t.department.department_id,
                }
                for t in teachers
            ]
            return Response({'pending': data}, status=status.HTTP_200_OK)

        if user_type == 'student':
            students = Student.objects.filter(verified=False)
            data = [
                {
                    'user_type': 'student',
                    'email': s.student_email,
                    'name': s.name,
                    'roll_no': s.roll_no,
                    'class_id': s.class_field.class_id,
                }
                for s in students
            ]
            return Response({'pending': data}, status=status.HTTP_200_OK)

        teachers = Teacher.objects.filter(verified=False)
        students = Student.objects.filter(verified=False)
        data = (
            [
                {
                    'user_type': 'teacher',
                    'email': t.teacher_email,
                    'name': t.name,
                    'designation': t.designation,
                    'department_id': t.department.department_id,
                }
                for t in teachers
            ]
            + [
                {
                    'user_type': 'student',
                    'email': s.student_email,
                    'name': s.name,
                    'roll_no': s.roll_no,
                    'class_id': s.class_field.class_id,
                }
                for s in students
            ]
        )

        return Response({'pending': data}, status=status.HTTP_200_OK)


class VerifyUserView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'admin':
            return Response({'error': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)

        serializer = VerifyUserSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user_type = serializer.validated_data['user_type']
        email = serializer.validated_data['email']
        verified = serializer.validated_data.get('verified', True)

        if user_type == 'teacher':
            teacher = Teacher.objects.filter(teacher_email=email).first()
            if not teacher:
                return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)
            teacher.verified = verified
            teacher.save()
            return Response({'message': 'Teacher updated'}, status=status.HTTP_200_OK)

        student = Student.objects.filter(student_email=email).first()
        if not student:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)
        student.verified = verified
        student.save()
        return Response({'message': 'Student updated'}, status=status.HTTP_200_OK)


class GoogleAuthView(APIView):
    """
    Google OAuth authentication with device security enforcement.

    SECURITY RULES FOR STUDENTS:
    1. Students MUST use mobile app (WEB platform blocked)
    2. Students can only be logged in on ONE device at a time
    3. New device requires teacher approval
    4. Logging in from a different device while another is active = BLOCKED

    TEACHERS:
    - No device restrictions, can login from any platform
    """
    permission_classes = []

    def post(self, request):
        serializer = GoogleAuthSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        id_token = data['id_token']
        device_uuid = data.get('device_uuid', '')
        fingerprint_hash = data.get('fingerprint_hash', '')
        platform = data.get('platform', 'WEB').upper()

        try:
            # Verify Firebase ID token
            decoded_token = firebase_auth.verify_id_token(id_token)
            email = decoded_token['email']
            name = decoded_token.get('name', email.split('@')[0])

        except Exception as e:
            return Response(
                {'error': f'Invalid token: {str(e)}'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Check if teacher - no device restrictions
        teacher = Teacher.objects.filter(teacher_email=email, verified=True).first()
        if teacher:
            user_info = {
                'email': teacher.teacher_email,
                'name': teacher.name,
                'user_type': 'teacher',
                'designation': teacher.designation,
                'department': teacher.department.department_name if teacher.department else None,
            }

            access_token = self._generate_jwt(teacher.teacher_email, 'teacher')
            refresh_token = self._generate_jwt(teacher.teacher_email, 'teacher', refresh=True)

            return Response({
                'access_token': access_token,
                'refresh_token': refresh_token,
                'user_type': 'teacher',
                'user_info': user_info,
                'device_approved': True,
                'device_id': None,
            })

        # Check if student - STRICT device security
        student = Student.objects.filter(student_email=email, verified=True).first()
        if student:
            # SECURITY: Students MUST use mobile app
            if platform == 'WEB':
                return Response(
                    {
                        'error': 'Students must use the mobile app to login',
                        'code': 'WEB_LOGIN_BLOCKED',
                        'message': 'For security reasons, students can only login through the official mobile app. Please download the app from the Play Store or App Store.'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )

            # SECURITY: Mobile must provide device info
            if not device_uuid or not fingerprint_hash:
                return Response(
                    {
                        'error': 'Device information required',
                        'code': 'DEVICE_INFO_MISSING',
                        'message': 'Please update your app to the latest version.'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Normalize platform
            if platform not in ['ANDROID', 'IOS']:
                platform = 'ANDROID'  # Default to Android for mobile

            # Check for existing active device
            active_device = Device.objects.filter(user_email=email, active=True).first()

            if active_device:
                # Check if it's the SAME device
                is_same_device = (
                    str(active_device.device_uuid) == device_uuid and
                    active_device.fingerprint_hash == fingerprint_hash
                )

                if is_same_device:
                    # Same device - allow login, update last_login
                    active_device.last_login = timezone.now()
                    active_device.save()

                    user_info = self._get_student_info(student)
                    access_token = self._generate_jwt(student.student_email, 'student', device_id=active_device.device_id)
                    refresh_token = self._generate_jwt(student.student_email, 'student', device_id=active_device.device_id, refresh=True)

                    return Response({
                        'access_token': access_token,
                        'refresh_token': refresh_token,
                        'user_type': 'student',
                        'user_info': user_info,
                        'device_approved': True,
                        'device_id': active_device.device_id,
                    })
                else:
                    # DIFFERENT device while another is active - BLOCK!
                    return Response(
                        {
                            'error': 'Already logged in on another device',
                            'code': 'DEVICE_CONFLICT',
                            'message': f'You are already logged in on another {active_device.platform} device. Please logout from that device first, or contact your teacher to reset your device.',
                            'active_device_platform': active_device.platform,
                            'active_device_registered': active_device.registered_at.isoformat(),
                        },
                        status=status.HTTP_403_FORBIDDEN
                    )

            # No active device - check if this device was previously registered
            existing_device = Device.objects.filter(
                user_email=email,
                device_uuid=device_uuid,
                fingerprint_hash=fingerprint_hash
            ).first()

            if existing_device:
                # Device exists but inactive - reactivate it
                # First, deactivate any other devices (shouldn't exist but safety check)
                Device.objects.filter(user_email=email).exclude(device_id=existing_device.device_id).update(active=False)

                existing_device.active = True
                existing_device.last_login = timezone.now()
                existing_device.platform = platform
                existing_device.save()

                user_info = self._get_student_info(student)
                access_token = self._generate_jwt(student.student_email, 'student', device_id=existing_device.device_id)
                refresh_token = self._generate_jwt(student.student_email, 'student', device_id=existing_device.device_id, refresh=True)

                return Response({
                    'access_token': access_token,
                    'refresh_token': refresh_token,
                    'user_type': 'student',
                    'user_info': user_info,
                    'device_approved': True,
                    'device_id': existing_device.device_id,
                })

            # NEW DEVICE - Register and require approval
            # Deactivate any existing devices first
            Device.objects.filter(user_email=email).update(active=False)

            new_device = Device.objects.create(
                user_email=email,
                device_uuid=device_uuid,
                fingerprint_hash=fingerprint_hash,
                integrity_level='BASIC',  # Can be enhanced with Play Integrity API
                platform=platform,
                active=True,  # Auto-approve first device for convenience
            )

            user_info = self._get_student_info(student)
            access_token = self._generate_jwt(student.student_email, 'student', device_id=new_device.device_id)
            refresh_token = self._generate_jwt(student.student_email, 'student', device_id=new_device.device_id, refresh=True)

            return Response({
                'access_token': access_token,
                'refresh_token': refresh_token,
                'user_type': 'student',
                'user_info': user_info,
                'device_approved': True,
                'device_id': new_device.device_id,
                'is_new_device': True,
            })

        return Response(
            {'error': f'User not found or not verified: {email}'},
            status=status.HTTP_404_NOT_FOUND
        )

    def _get_student_info(self, student):
        return {
            'email': student.student_email,
            'name': student.name,
            'user_type': 'student',
            'roll_no': student.roll_no,
            'class_id': student.class_field.class_id if student.class_field else None,
        }

    def _generate_jwt(self, email, user_type, device_id=None, refresh=False):
        return _encode_jwt(email, user_type, device_id=device_id, refresh=refresh)

class ApproveDeviceView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response(
                {'error': 'Only teachers can approve devices'},
                status=status.HTTP_403_FORBIDDEN
            )

        student_email = request.data.get('student_email')
        device_id = request.data.get('device_id')

        if not student_email or not device_id:
            return Response(
                {'error': 'student_email and device_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        student = Student.objects.filter(student_email=student_email, verified=True).first()
        if not student:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        device = Device.objects.filter(device_id=device_id, user_email=student_email).first()
        if not device:
            return Response({'error': 'Device not found'}, status=status.HTTP_404_NOT_FOUND)

        Device.objects.filter(user_email=student_email, active=True).update(active=False)
        device.active = True
        device.last_login = timezone.now()
        device.save()

        return Response({'message': 'Device approved'}, status=status.HTTP_200_OK)


class StudentLogoutView(APIView):
    """
    Logout endpoint for students - deactivates their current device.
    This allows them to login from a different device.
    """
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'student':
            return Response(
                {'error': 'This endpoint is for students only'},
                status=status.HTTP_403_FORBIDDEN
            )

        email = str(request.user)

        # Deactivate all devices for this student
        devices_deactivated = Device.objects.filter(user_email=email, active=True).update(active=False)

        return Response({
            'message': 'Logged out successfully',
            'devices_deactivated': devices_deactivated
        }, status=status.HTTP_200_OK)


class ResetStudentDeviceView(APIView):
    """
    Teacher endpoint to reset a student's device.
    Use this when a student needs to switch devices or is locked out.
    """
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response(
                {'error': 'Only teachers can reset student devices'},
                status=status.HTTP_403_FORBIDDEN
            )

        student_email = request.data.get('student_email')
        if not student_email:
            return Response(
                {'error': 'student_email required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        student = Student.objects.filter(student_email=student_email).first()
        if not student:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Verify teacher has authority over this student (same class)
        teacher_email = str(request.user)
        teacher = Teacher.objects.filter(teacher_email=teacher_email).first()

        # Check if teacher teaches any class the student is in
        student_class_ids = list(StudentClass.objects.filter(student=student).values_list('class_obj_id', flat=True))
        if student.class_field_id:
            student_class_ids.append(student.class_field_id)

        teacher_class_ids = list(Subject.objects.filter(teacher_email=teacher_email).values_list('class_field_id', flat=True))
        teacher_class_ids += list(ClassTeacher.objects.filter(teacher=teacher).values_list('class_obj_id', flat=True))

        has_authority = bool(set(student_class_ids) & set(teacher_class_ids))
        if not has_authority:
            return Response(
                {'error': 'You do not have authority over this student'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Deactivate and delete all devices for this student
        devices_deleted = Device.objects.filter(user_email=student_email).delete()[0]

        return Response({
            'message': f'Device reset successful for {student.name}',
            'devices_removed': devices_deleted,
            'student_can_login': True,
        }, status=status.HTTP_200_OK)


class StudentDeviceInfoView(APIView):
    """
    Get device information for a student (teacher endpoint).
    """
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, email):
        if request.user_type != 'teacher':
            return Response(
                {'error': 'Only teachers can view device info'},
                status=status.HTTP_403_FORBIDDEN
            )

        devices = Device.objects.filter(user_email=email).values(
            'device_id', 'platform', 'active', 'registered_at', 'last_login'
        )

        return Response({
            'student_email': email,
            'devices': list(devices)
        })


@method_decorator(csrf_exempt, name='dispatch')
class InitiateAttendanceView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user

        # Accept either period_id OR subject_id
        period_id = request.data.get('period_id')
        subject_id = request.data.get('subject_id')
        date_str = request.data.get('date', str(timezone.now().date()))

        # GPS proximity fields
        class_mode = request.data.get('class_mode', 'offline')
        teacher_latitude = request.data.get('teacher_latitude')
        teacher_longitude = request.data.get('teacher_longitude')
        proximity_radius = request.data.get('proximity_radius', 30)

        if not period_id and not subject_id:
            return Response(
                {'error': 'Either period_id or subject_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get teacher
            teacher = Teacher.objects.get(teacher_email=teacher_email, verified=True)

            if period_id:
                # Get period directly
                period = Period.objects.select_related(
                    'subject',
                    'subject__class_field',
                    'subject__course'
                ).get(period_id=period_id)
                subject = period.subject
            else:
                # Get subject and find/create period for today
                subject = Subject.objects.select_related(
                    'class_field', 'course'
                ).get(subject_id=subject_id)

                today_weekday = timezone.now().weekday()

                # Try to find existing period for today
                period = Period.objects.filter(
                    subject=subject,
                    day_of_week=today_weekday
                ).first()

                # If no period exists, create an ad-hoc one
                if not period:
                    # Find the next available period number for this day
                    max_period = Period.objects.filter(
                        subject__class_field=subject.class_field,
                        day_of_week=today_weekday
                    ).aggregate(max_no=Max('period_no'))['max_no'] or 0

                    period = Period.objects.create(
                        subject=subject,
                        day_of_week=today_weekday,
                        period_no=max_period + 1
                    )

            # Verify teacher owns this subject
            if subject.teacher.teacher_email != teacher_email:
                return Response(
                    {'error': 'You do not teach this subject'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Check if class is active (semester not completed)
            if not subject.class_field.is_active:
                return Response(
                    {'error': 'Cannot take attendance - semester has been marked as complete'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Check for existing ACTIVE session for this period on this date
            active_session = Session.objects.filter(
                period=period,
                date=date_str,
                is_active=True
            ).first()

            if active_session:
                # Return the existing active session instead of creating new one
                return Response({
                    'error': 'Active session already exists for this subject today',
                    'session_id': active_session.session_id,
                    'otp': active_session.otp
                }, status=status.HTTP_400_BAD_REQUEST)

            # Generate alphanumeric OTP and create NEW session (allows multiple sessions per day)
            otp = generate_alphanumeric_otp(4)
            session = Session.objects.create(
                period=period,
                date=date_str,
                otp=otp,
                otp_generated_at=timezone.now(),
                is_active=True,
                class_mode=class_mode if class_mode in ('offline', 'online') else 'offline',
                teacher_latitude=teacher_latitude if class_mode == 'offline' else None,
                teacher_longitude=teacher_longitude if class_mode == 'offline' else None,
                proximity_radius=int(proximity_radius) if proximity_radius else 30,
            )

            # Get all students in this class (from both class_field AND StudentClass enrollments)
            # 1. Students with this as primary class
            primary_students = Student.objects.filter(
                class_field=subject.class_field,
                verified=True
            )

            # 2. Students enrolled via StudentClass
            enrolled_student_emails = StudentClass.objects.filter(
                class_obj=subject.class_field
            ).values_list('student__student_email', flat=True)

            enrolled_students = Student.objects.filter(
                student_email__in=enrolled_student_emails,
                verified=True
            )

            # Combine and deduplicate
            all_student_emails = set(primary_students.values_list('student_email', flat=True)) | set(enrolled_student_emails)
            students = Student.objects.filter(
                student_email__in=all_student_emails,
                verified=True
            ).order_by('roll_no')

            # Create attendance records for all students (default: Absent)
            attendance_records = []
            for student in students:
                attendance_records.append(
                    Attendance(
                        session=session,
                        student=student,
                        status='A',
                        submitted_at=None
                    )
                )

            if attendance_records:
                Attendance.objects.bulk_create(attendance_records, ignore_conflicts=True)

            # Get student emails
            student_emails = list(students.values_list('student_email', flat=True))

            # Broadcast to all students
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                'all_students',
                {
                    'type': 'attendance_started',
                    'session_id': session.session_id,
                    'message': 'Attendance session started'
                }
            )

            return Response({
                'session_id': session.session_id,
                'otp': otp,
                'expires_in': getattr(settings, 'OTP_VALIDITY_SECONDS', 30),
                'enrolled_students': student_emails,
                'total_students': len(student_emails),
                'class_name': str(subject.class_field),
                'subject_name': subject.course.course_name,
                'class_mode': session.class_mode,
            }, status=status.HTTP_201_CREATED)

        except Teacher.DoesNotExist:
            return Response(
                {'error': 'Teacher not found or not verified'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Period.DoesNotExist:
            return Response(
                {'error': 'Period not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Subject.DoesNotExist:
            return Response(
                {'error': 'Subject not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@method_decorator(csrf_exempt, name='dispatch')
class CloseAttendanceView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user
        session_id = request.data.get('session_id')

        try:
            # Get session and verify teacher owns it
            session = Session.objects.select_related(
                'period__subject__teacher'
            ).get(session_id=session_id)

            if session.period.subject.teacher.teacher_email != teacher_email:
                return Response(
                    {'error': 'Unauthorized'},
                    status=status.HTTP_403_FORBIDDEN
                )

            if not session.is_active:
                return Response(
                    {'error': 'Session already closed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Mark all pending (not submitted) students as absent
            Attendance.objects.filter(
                session=session,
                submitted_at__isnull=True
            ).update(
                status='A',
                submitted_at=timezone.now()
            )

            # Mark all Proxy students as Absent (proxy = out of GPS range)
            Attendance.objects.filter(
                session=session,
                status='X'
            ).update(status='A')

            # Close session
            session.is_active = False
            session.closed_at = timezone.now()
            session.save()

            # Broadcast
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                'all_students',
                {
                    'type': 'attendance_closed',
                    'session_id': session_id,
                    'message': 'Session closed'
                }
            )

            return Response(
                {'message': 'Session closed successfully'},
                status=status.HTTP_200_OK
            )

        except Session.DoesNotExist:
            return Response(
                {'error': 'Session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@method_decorator(csrf_exempt, name='dispatch')
class RegenerateOTPView(APIView):
    """Regenerate OTP for an active session (manual regeneration by teacher)"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user
        session_id = request.data.get('session_id')

        if not session_id:
            return Response(
                {'error': 'session_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get session and verify teacher owns it
            session = Session.objects.select_related(
                'period__subject__teacher'
            ).get(session_id=session_id)

            if session.period.subject.teacher.teacher_email != teacher_email:
                return Response(
                    {'error': 'Unauthorized'},
                    status=status.HTTP_403_FORBIDDEN
                )

            if not session.is_active:
                return Response(
                    {'error': 'Session is closed'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Generate new alphanumeric OTP
            new_otp = generate_alphanumeric_otp(4)
            session.otp = new_otp
            session.otp_generated_at = timezone.now()
            session.save()

            # Broadcast OTP regeneration to all students
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                'all_students',
                {
                    'type': 'otp_regenerated',
                    'session_id': session_id,
                    'message': 'New OTP generated'
                }
            )

            return Response({
                'session_id': session.session_id,
                'otp': new_otp,
                'expires_in': getattr(settings, 'OTP_VALIDITY_SECONDS', 30),
                'message': 'OTP regenerated successfully'
            }, status=status.HTTP_200_OK)

        except Session.DoesNotExist:
            return Response(
                {'error': 'Session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@method_decorator(csrf_exempt, name='dispatch')
class LiveStatusView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user
        session_id = request.query_params.get('session_id')

        if not session_id:
            return Response(
                {'error': 'session_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get session
            session = Session.objects.select_related(
                'period__subject__teacher'
            ).get(session_id=session_id)

            if session.period.subject.teacher.teacher_email != teacher_email:
                return Response(
                    {'error': 'Unauthorized'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Get all attendance records
            attendances = Attendance.objects.filter(
                session=session
            ).select_related('student').order_by('student__roll_no')

            total = attendances.count()
            present = attendances.filter(status='P').count()
            proxy = attendances.filter(status='X').count()

            # If session is active, differentiate between pending and absent
            if session.is_active:
                pending = attendances.filter(submitted_at__isnull=True).count()
                absent = attendances.filter(status='A', submitted_at__isnull=False).count()
            else:
                pending = 0
                absent = attendances.filter(status='A').count()

            # Get roll numbers from StudentClass enrollments for this class
            class_obj = session.period.subject.class_field
            enrollment_roll_nos = {
                sc.student.student_email: sc.roll_no
                for sc in StudentClass.objects.filter(class_obj=class_obj).select_related('student')
            }

            # Build submissions list
            submissions = []
            for att in attendances:
                # Prefer roll_no from enrollment, fallback to student's default
                roll_no = enrollment_roll_nos.get(att.student.student_email, att.student.roll_no)
                status_display_map = {'P': 'Present', 'A': 'Absent', 'X': 'Proxy', 'R': 'Retry'}
                submissions.append({
                    'student_email': att.student.student_email,
                    'student_name': att.student.name,
                    'roll_no': roll_no,
                    'status': att.status,
                    'status_display': status_display_map.get(att.status, 'Absent'),
                    'submitted_at': att.submitted_at.isoformat() if att.submitted_at else None
                })

            return Response({
                'session_id': session.session_id,
                'total_students': total,
                'present': present,
                'absent': absent,
                'pending': pending,
                'proxy': proxy,
                'class_mode': session.class_mode,
                'submissions': submissions
            }, status=status.HTTP_200_OK)

        except Session.DoesNotExist:
            return Response(
                {'error': 'Session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@method_decorator(csrf_exempt, name='dispatch')
class ManualMarkView(APIView):
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user
        session_id = request.data.get('session_id')
        student_email = request.data.get('student_email')
        new_status = request.data.get('status')  # 'P' or 'A'

        if not all([session_id, student_email, new_status]):
            return Response(
                {'error': 'session_id, student_email, and status required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if new_status not in ['P', 'A']:
            return Response(
                {'error': 'status must be P or A'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Get session and verify teacher owns it
            session = Session.objects.select_related(
                'period__subject__teacher'
            ).get(session_id=session_id)

            if session.period.subject.teacher.teacher_email != teacher_email:
                return Response(
                    {'error': 'Unauthorized'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Get attendance record
            attendance = Attendance.objects.get(
                session=session,
                student__student_email=student_email
            )

            # Update status
            attendance.status = new_status
            if not attendance.submitted_at:
                attendance.submitted_at = timezone.now()
            attendance.save()

            # Broadcast update
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                'all_teachers',
                {
                    'type': 'attendance_update',
                    'session_id': session.session_id,
                    'student_email': student_email,
                    'status': new_status
                }
            )

            return Response(
                {'message': 'Attendance updated successfully'},
                status=status.HTTP_200_OK
            )

        except Session.DoesNotExist:
            return Response(
                {'error': 'Session not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Attendance.DoesNotExist:
            return Response(
                {'error': 'Attendance record not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class StudentSummaryView(APIView):
    """
    Get attendance summary for students
    """
    permission_classes = [IsJWTAuthenticated]
    
    def get(self, request):
        student_email = request.query_params.get('student_email')
        subject_id = request.query_params.get('subject_id')
        
        if not student_email or not subject_id:
            return Response(
                {'error': 'student_email and subject_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        student = Student.objects.filter(student_email=student_email).first()
        if not student:
            return Response(
                {'error': 'Student not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get all sessions for this subject
        sessions = Session.objects.filter(
            period__subject__subject_id=subject_id
        )
        
        attendances = Attendance.objects.filter(
            student=student,
            session__in=sessions
        )
        
        total = attendances.count()
        present = attendances.filter(status='P').count()
        absent = attendances.filter(status='A').count()
        
        percentage = (present / total * 100) if total > 0 else 0
        
        response_data = {
            'student_email': student_email,
            'student_name': student.name,
            'roll_no': student.roll_no,
            'total_sessions': total,
            'present': present,
            'absent': absent,
            'attendance_percentage': round(percentage, 2)
        }
        
        return Response(response_data, status=status.HTTP_200_OK)


class ClassSummaryView(APIView):
    """
    Get attendance summary for entire class
    """
    permission_classes = [IsJWTAuthenticated]
    
    def get(self, request):
        if request.user_type != 'teacher':
            return Response(
                {'error': 'Only teachers can view class summary'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        subject_id = request.query_params.get('subject_id')
        
        if not subject_id:
            return Response(
                {'error': 'subject_id required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        subject = Subject.objects.filter(
            subject_id=subject_id,
            teacher__teacher_email=request.user
        ).first()
        
        if not subject:
            return Response(
                {'error': 'Subject not found or unauthorized'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get all students in class (from both class_field AND StudentClass enrollments)
        primary_students = Student.objects.filter(class_field=subject.class_field)
        enrolled_student_emails = StudentClass.objects.filter(
            class_obj=subject.class_field
        ).values_list('student__student_email', flat=True)

        all_student_emails = set(primary_students.values_list('student_email', flat=True)) | set(enrolled_student_emails)
        students = Student.objects.filter(student_email__in=all_student_emails)

        # Get roll numbers from enrollments
        enrollment_roll_nos = {
            sc.student.student_email: sc.roll_no
            for sc in StudentClass.objects.filter(class_obj=subject.class_field).select_related('student')
        }

        # Get all sessions for this subject
        sessions = Session.objects.filter(period__subject=subject)
        total_sessions = sessions.count()

        student_summaries = []
        total_attendance = 0

        for student in students:
            attendances = Attendance.objects.filter(
                student=student,
                session__in=sessions
            )

            present = attendances.filter(status='P').count()
            absent = attendances.filter(status='A').count()
            percentage = (present / total_sessions * 100) if total_sessions > 0 else 0

            total_attendance += percentage

            # Prefer roll_no from enrollment
            roll_no = enrollment_roll_nos.get(student.student_email, student.roll_no)

            student_summaries.append({
                'student_email': student.student_email,
                'student_name': student.name,
                'roll_no': roll_no,
                'total_sessions': total_sessions,
                'present': present,
                'absent': absent,
                'attendance_percentage': round(percentage, 2)
            })

        avg_attendance = (total_attendance / len(students)) if students else 0
        
        response_data = {
            'class_id': subject.class_field.class_id,
            'class_info': str(subject.class_field),
            'subject_name': subject.course.course_name,
            'total_sessions': total_sessions,
            'average_attendance': round(avg_attendance, 2),
            'students': student_summaries
        }

        return Response(response_data, status=status.HTTP_200_OK)


# ============== CLASS MANAGEMENT APIs ==============

class ClassListCreateView(APIView):
    """List and create classes for a teacher"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        # Get classes where teacher is:
        # 1. Coordinator
        # 2. Creator
        # 3. Assigned teacher (via ClassTeacher)
        # 4. Has subjects to teach
        teacher_email = request.user

        coordinated_classes = ClassModel.objects.filter(coordinator__teacher_email=teacher_email)
        created_classes = ClassModel.objects.filter(created_by__teacher_email=teacher_email)
        assigned_classes = ClassModel.objects.filter(assigned_teachers__teacher__teacher_email=teacher_email)
        taught_classes = ClassModel.objects.filter(
            subject__teacher__teacher_email=teacher_email
        ).distinct()

        # Combine and deduplicate
        class_ids = (
            set(coordinated_classes.values_list('class_id', flat=True)) |
            set(created_classes.values_list('class_id', flat=True)) |
            set(assigned_classes.values_list('class_id', flat=True)) |
            set(taught_classes.values_list('class_id', flat=True))
        )
        classes = ClassModel.objects.filter(class_id__in=class_ids).select_related(
            'department', 'created_by', 'coordinator'
        )

        # Build response with role information
        result = []
        for cls in classes:
            data = ClassDetailSerializer(cls).data
            # Add coordinator info
            data['coordinator_email'] = cls.coordinator.teacher_email if cls.coordinator else None
            data['coordinator_name'] = cls.coordinator.name if cls.coordinator else None
            data['is_coordinator'] = cls.coordinator and cls.coordinator.teacher_email == teacher_email
            result.append(data)

        return Response(result)

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        serializer = ClassCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        teacher = Teacher.objects.get(teacher_email=request.user)

        # Check if class already exists
        existing = ClassModel.objects.filter(
            department=serializer.validated_data['department'],
            batch=serializer.validated_data['batch'],
            semester=serializer.validated_data['semester'],
            section=serializer.validated_data['section']
        ).first()

        if existing:
            return Response({'error': 'Class already exists'}, status=status.HTTP_400_BAD_REQUEST)

        # Creator becomes the coordinator
        class_obj = ClassModel.objects.create(
            **serializer.validated_data,
            created_by=teacher,
            coordinator=teacher
        )

        data = ClassDetailSerializer(class_obj).data
        data['coordinator_email'] = teacher.teacher_email
        data['coordinator_name'] = teacher.name
        data['is_coordinator'] = True

        return Response(data, status=status.HTTP_201_CREATED)


class ClassDetailView(APIView):
    """Get, update, delete a specific class"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('department', 'created_by').get(class_id=pk)
            return Response(ClassDetailSerializer(class_obj).data)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk, created_by__teacher_email=request.user)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found or unauthorized'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ClassCreateSerializer(class_obj, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response(ClassDetailSerializer(class_obj).data)

    def delete(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('coordinator').get(class_id=pk)

            # Only coordinator can delete the class
            is_coordinator = class_obj.coordinator and class_obj.coordinator.teacher_email == str(request.user)

            if not is_coordinator:
                return Response({'error': 'Unauthorized - only the class coordinator can delete this class'}, status=status.HTTP_403_FORBIDDEN)

            class_obj.delete()
            return Response({'message': 'Class deleted'}, status=status.HTTP_204_NO_CONTENT)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)


class ClassStudentsView(APIView):
    """List students in a class"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get students from StudentClass (enrolled students)
        enrollments = StudentClass.objects.filter(class_obj=class_obj).select_related('student')
        enrollment_data = StudentEnrollmentSerializer(enrollments, many=True).data

        # Also get students with this as primary class
        primary_students = Student.objects.filter(class_field=class_obj).exclude(
            student_email__in=[e['student_email'] for e in enrollment_data]
        )

        for student in primary_students:
            enrollment_data.append({
                'id': None,
                'student_email': student.student_email,
                'student_name': student.name,
                'roll_no': student.roll_no,
                'enrolled_at': None,
                'verified': student.verified
            })

        return Response({
            'class_id': pk,
            'class_name': str(class_obj),
            'students': enrollment_data
        })


# ============== STUDENT LOOKUP API ==============

class StudentLookupView(APIView):
    """Lookup a student by email - returns name and roll_no if exists"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        email = request.query_params.get('email', '').strip().lower()
        if not email:
            return Response({'error': 'email parameter required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if student exists
        student = Student.objects.filter(student_email=email).first()

        if student:
            return Response({
                'exists': True,
                'student_email': student.student_email,
                'name': student.name,
                'roll_no': student.roll_no,
                'verified': student.verified
            })

        # Check if there's a pending invitation for this email
        invitation = StudentInvitation.objects.filter(email=email).first()
        if invitation:
            return Response({
                'exists': False,
                'pending_invitation': True,
                'name': invitation.name,
                'roll_no': invitation.roll_no
            })

        return Response({
            'exists': False,
            'pending_invitation': False
        })


# ============== STUDENT INVITATION APIs ==============

class InviteStudentView(APIView):
    """Add a student to a class (create if doesn't exist)"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = StudentInviteSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        email = serializer.validated_data['email']
        name = serializer.validated_data.get('name', '')
        roll_no = serializer.validated_data['roll_no']

        # Check if student already exists
        student = Student.objects.filter(student_email=email).first()

        if student:
            # Check if already enrolled in this class
            if StudentClass.objects.filter(student=student, class_obj=class_obj).exists():
                return Response({'error': 'Student already enrolled in this class'}, status=status.HTTP_400_BAD_REQUEST)

            # Update student name if provided (this persists globally)
            if name and name.strip():
                student.name = name.strip()
                student.save()

            # Enroll the student
            StudentClass.objects.create(
                student=student,
                class_obj=class_obj,
                roll_no=roll_no
            )
            return Response({'message': 'Student enrolled successfully', 'enrolled': True})

        # Student doesn't exist - create them directly
        # Use the name if provided, otherwise use email prefix
        student_name = name.strip() if name and name.strip() else email.split('@')[0]

        student = Student.objects.create(
            student_email=email,
            name=student_name,
            roll_no=roll_no,
            class_field=class_obj,
            verified=True  # Mark as verified so they can log in
        )

        # Also enroll in the class via StudentClass
        StudentClass.objects.create(
            student=student,
            class_obj=class_obj,
            roll_no=roll_no
        )

        return Response({
            'message': 'Student added successfully',
            'enrolled': True,
            'created': True
        }, status=status.HTTP_201_CREATED)


class ImportStudentsView(APIView):
    """Bulk import students from CSV or Excel file"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'error': 'File required (CSV or Excel)'}, status=status.HTTP_400_BAD_REQUEST)

        teacher = Teacher.objects.get(teacher_email=request.user)
        filename = uploaded_file.name.lower()

        try:
            rows = []

            # Handle Excel files
            if filename.endswith(('.xlsx', '.xls')):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=uploaded_file, read_only=True)
                    ws = wb.active

                    # Get header row
                    headers = []
                    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                        headers.append(str(cell.value).lower().strip() if cell.value else '')

                    # Process data rows
                    for row in ws.iter_rows(min_row=2):
                        row_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = str(cell.value).strip() if cell.value else ''
                        if any(row_data.values()):  # Skip empty rows
                            rows.append(row_data)
                    wb.close()
                except ImportError:
                    return Response(
                        {'error': 'Excel support not installed. Please use CSV format.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            # Handle CSV files
            elif filename.endswith('.csv'):
                decoded_file = uploaded_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(reader)
            else:
                return Response(
                    {'error': 'Unsupported file format. Please use CSV or Excel (.xlsx, .xls)'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            results = {'enrolled': 0, 'created': 0, 'errors': []}

            for row in rows:
                email = row.get('email', '').strip()
                name = row.get('name', '').strip()
                roll_no = row.get('roll_no', '').strip()

                if not email or not roll_no:
                    results['errors'].append(f"Missing email or roll_no: {row}")
                    continue

                student = Student.objects.filter(student_email=email).first()

                if student:
                    if not StudentClass.objects.filter(student=student, class_obj=class_obj).exists():
                        # Update name if provided
                        if name:
                            student.name = name
                            student.save()
                        StudentClass.objects.create(
                            student=student,
                            class_obj=class_obj,
                            roll_no=roll_no
                        )
                        results['enrolled'] += 1
                else:
                    # Create student directly
                    student_name = name if name else email.split('@')[0]
                    student = Student.objects.create(
                        student_email=email,
                        name=student_name,
                        roll_no=roll_no,
                        class_field=class_obj,
                        verified=True  # Mark as verified so they can log in
                    )
                    # Also enroll in the class via StudentClass
                    StudentClass.objects.create(
                        student=student,
                        class_obj=class_obj,
                        roll_no=roll_no
                    )
                    results['created'] += 1

            return Response(results)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class RemoveStudentView(APIView):
    """Remove a student from a class"""
    permission_classes = [IsJWTAuthenticated]

    def delete(self, request, pk, email):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Remove from StudentClass
        deleted, _ = StudentClass.objects.filter(
            class_obj=class_obj,
            student__student_email=email
        ).delete()

        if deleted:
            return Response({'message': 'Student removed from class'})

        return Response({'error': 'Student not found in class'}, status=status.HTTP_404_NOT_FOUND)


# ============== SCHEDULE/PERIOD APIs ==============

class ScheduleListView(APIView):
    """Get teacher's weekly schedule, or all periods for a class if class_id provided"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = str(request.user)
        day = request.query_params.get('day')
        class_id = request.query_params.get('class_id')

        if class_id:
            # Return ALL periods for this class (for schedule viewing)
            # Include ownership flags so UI knows what's editable
            periods = Period.objects.filter(
                subject__class_field_id=class_id
            ).select_related(
                'subject__course', 'subject__class_field',
                'subject__class_field__department', 'subject__class_field__coordinator',
                'subject__teacher'
            )
        else:
            # Default: only teacher's own subjects
            periods = Period.objects.filter(
                subject__teacher__teacher_email=teacher_email
            ).select_related('subject__course', 'subject__class_field', 'subject__class_field__department')

        if day is not None:
            periods = periods.filter(day_of_week=int(day))

        periods = periods.order_by('day_of_week', 'period_no')

        schedule = []
        for period in periods:
            is_subject_owner = period.subject.teacher.teacher_email == teacher_email
            is_coordinator = (
                period.subject.class_field.coordinator and
                period.subject.class_field.coordinator.teacher_email == teacher_email
            )

            schedule.append({
                'period_id': period.period_id,
                'subject_id': period.subject.subject_id,
                'subject_name': f"{period.subject.course.course_name} - {period.subject.class_field}",
                'course_name': period.subject.course.course_name,
                'teacher_name': period.subject.teacher.name,
                'class_name': str(period.subject.class_field),
                'class_id': period.subject.class_field.class_id,
                'day_of_week': period.day_of_week,
                'day_name': period.get_day_of_week_display(),
                'period_no': period.period_no,
                'is_subject_owner': is_subject_owner,
                'is_coordinator': is_coordinator,
                'can_manage': is_subject_owner or is_coordinator
            })

        return Response(schedule)


class TodayScheduleView(APIView):
    """Get today's schedule for teacher"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        today = timezone.now().weekday()  # Monday = 0

        periods = Period.objects.filter(
            subject__teacher__teacher_email=request.user,
            day_of_week=today
        ).select_related('subject__course', 'subject__class_field').order_by('period_no')

        schedule = []
        for period in periods:
            schedule.append({
                'period_id': period.period_id,
                'subject_id': period.subject.subject_id,
                'subject_name': f"{period.subject.course.course_name} - {period.subject.class_field}",
                'course_name': period.subject.course.course_name,
                'class_name': str(period.subject.class_field),
                'class_id': period.subject.class_field.class_id,
                'period_no': period.period_no
            })

        return Response(schedule)


class PeriodCreateView(APIView):
    """Create a new period"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        serializer = PeriodCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        teacher_email = str(request.user)
        subject_id = serializer.validated_data['subject'].subject_id

        # Get the subject with class coordinator info
        subject = Subject.objects.select_related('class_field__coordinator').filter(
            subject_id=subject_id
        ).first()

        if not subject:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if teacher is subject owner OR class coordinator
        is_subject_owner = subject.teacher.teacher_email == teacher_email
        is_coordinator = subject.class_field.coordinator and subject.class_field.coordinator.teacher_email == teacher_email

        if not (is_subject_owner or is_coordinator):
            return Response({'error': 'Unauthorized - only subject teacher or class coordinator can add periods'}, status=status.HTTP_403_FORBIDDEN)

        # Check for existing period at same time
        existing = Period.objects.filter(
            subject=subject,
            day_of_week=serializer.validated_data['day_of_week'],
            period_no=serializer.validated_data['period_no']
        ).first()

        if existing:
            return Response({'error': 'Period already exists at this time'}, status=status.HTTP_400_BAD_REQUEST)

        period = Period.objects.create(**serializer.validated_data)
        return Response(PeriodSerializer(period).data, status=status.HTTP_201_CREATED)


class PeriodDetailView(APIView):
    """Update or delete a period"""
    permission_classes = [IsJWTAuthenticated]

    def _can_manage_period(self, period, teacher_email):
        """Check if teacher can manage this period (is subject owner or coordinator)"""
        is_subject_owner = period.subject.teacher.teacher_email == teacher_email
        is_coordinator = (
            period.subject.class_field.coordinator and
            period.subject.class_field.coordinator.teacher_email == teacher_email
        )
        return is_subject_owner or is_coordinator

    def put(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            period = Period.objects.select_related(
                'subject__teacher', 'subject__class_field__coordinator'
            ).get(period_id=pk)
        except Period.DoesNotExist:
            return Response({'error': 'Period not found'}, status=status.HTTP_404_NOT_FOUND)

        if not self._can_manage_period(period, str(request.user)):
            return Response({'error': 'Unauthorized - only subject teacher or class coordinator can update periods'}, status=status.HTTP_403_FORBIDDEN)

        serializer = PeriodCreateSerializer(period, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response(PeriodSerializer(period).data)

    def delete(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            period = Period.objects.select_related(
                'subject__teacher', 'subject__class_field__coordinator'
            ).get(period_id=pk)
        except Period.DoesNotExist:
            return Response({'error': 'Period not found'}, status=status.HTTP_404_NOT_FOUND)

        if not self._can_manage_period(period, str(request.user)):
            return Response({'error': 'Unauthorized - only subject teacher or class coordinator can delete periods'}, status=status.HTTP_403_FORBIDDEN)

        period.delete()
        return Response({'message': 'Period deleted'}, status=status.HTTP_204_NO_CONTENT)


class SubjectListCreateView(APIView):
    """List and create subjects for a teacher"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = str(request.user)
        class_id = request.query_params.get('class_id')

        if class_id:
            # If class_id provided, return all subjects for that class with ownership info
            subjects = Subject.objects.filter(
                class_field_id=class_id
            ).select_related('course', 'class_field', 'class_field__department', 'class_field__coordinator', 'teacher')

            result = []
            for subject in subjects:
                data = SubjectSerializer(subject).data
                data['is_subject_owner'] = subject.teacher.teacher_email == teacher_email
                data['is_coordinator'] = (
                    subject.class_field.coordinator and
                    subject.class_field.coordinator.teacher_email == teacher_email
                )
                data['can_manage'] = data['is_subject_owner'] or data['is_coordinator']
                result.append(data)

            return Response(result)
        else:
            # Default: return subjects taught by this teacher
            subjects = Subject.objects.filter(
                teacher__teacher_email=teacher_email
            ).select_related('course', 'class_field', 'class_field__department')

            result = []
            for subject in subjects:
                data = SubjectSerializer(subject).data
                data['is_subject_owner'] = True  # Teacher's own subjects
                data['can_manage'] = True
                result.append(data)

            return Response(result)

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher = Teacher.objects.get(teacher_email=request.user)

        # Handle course - create if doesn't exist
        course_name = request.data.get('course_name')
        course_id = request.data.get('course')

        if course_name:
            course, _ = Course.objects.get_or_create(course_name=course_name)
        elif course_id:
            try:
                course = Course.objects.get(course_id=course_id)
            except Course.DoesNotExist:
                return Response({'error': 'Course not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'error': 'course or course_name required'}, status=status.HTTP_400_BAD_REQUEST)

        class_id = request.data.get('class_field')
        try:
            class_obj = ClassModel.objects.get(class_id=class_id)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if subject already exists
        existing = Subject.objects.filter(course=course, class_field=class_obj).first()
        if existing:
            return Response({'error': 'Subject already exists for this class'}, status=status.HTTP_400_BAD_REQUEST)

        subject = Subject.objects.create(
            course=course,
            class_field=class_obj,
            teacher=teacher
        )

        return Response(SubjectSerializer(subject).data, status=status.HTTP_201_CREATED)


class CourseListView(APIView):
    """List all courses"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        courses = Course.objects.all()
        return Response(CourseSerializer(courses, many=True).data)


class DepartmentListView(APIView):
    """List departments - filtered by teacher's organization if authenticated as teacher"""
    permission_classes = []  # Allow any access - departments are not sensitive

    def get(self, request):
        # If authenticated as teacher, filter by their organization
        if hasattr(request, 'user_type') and request.user_type == 'teacher':
            try:
                teacher = Teacher.objects.get(teacher_email=request.user)
                if teacher.organization:
                    departments = Department.objects.filter(organization=teacher.organization)
                else:
                    departments = Department.objects.all()
            except Teacher.DoesNotExist:
                departments = Department.objects.all()
        else:
            departments = Department.objects.all()
        return Response(DepartmentSerializer(departments, many=True).data)


class OrganizationListView(APIView):
    """List all organizations"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        from .models import Organization
        from .serializers import OrganizationSerializer
        organizations = Organization.objects.all()
        return Response(OrganizationSerializer(organizations, many=True).data)


class MarkSemesterCompleteView(APIView):
    """Mark a class/semester as complete - no more attendance can be taken"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('coordinator').get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        try:
            teacher = Teacher.objects.get(teacher_email=request.user)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

        # Only coordinator can mark semester as complete
        is_coordinator = class_obj.coordinator and class_obj.coordinator.teacher_email == str(request.user)

        if not is_coordinator:
            return Response({'error': 'Unauthorized - only the class coordinator can mark semester as complete'},
                          status=status.HTTP_403_FORBIDDEN)

        if not class_obj.is_active:
            return Response({'error': 'Class already completed'}, status=status.HTTP_400_BAD_REQUEST)

        # Close any active sessions for this class
        Session.objects.filter(
            period__subject__class_field=class_obj,
            is_active=True
        ).update(is_active=False, closed_at=timezone.now())

        # Mark class as complete
        class_obj.is_active = False
        class_obj.completed_at = timezone.now()
        class_obj.completed_by = teacher
        class_obj.save()

        return Response({
            'message': 'Semester marked as complete',
            'class_id': class_obj.class_id,
            'completed_at': class_obj.completed_at
        })


class TeacherClassesForAttendanceView(APIView):
    """Get teacher's active classes with subjects for attendance selection"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        subjects = Subject.objects.filter(
            teacher__teacher_email=request.user,
            class_field__is_active=True  # Only active classes
        ).select_related('course', 'class_field', 'class_field__department')

        result = []
        for subject in subjects:
            # Count students from both class_field AND StudentClass enrollments
            primary_students = Student.objects.filter(
                class_field=subject.class_field,
                verified=True
            ).values_list('student_email', flat=True)

            enrolled_students = StudentClass.objects.filter(
                class_obj=subject.class_field,
                student__verified=True
            ).values_list('student__student_email', flat=True)

            all_student_emails = set(primary_students) | set(enrolled_students)
            student_count = len(all_student_emails)

            result.append({
                'subject_id': subject.subject_id,
                'subject_name': subject.course.course_name,
                'class_id': subject.class_field.class_id,
                'class_name': str(subject.class_field),
                'department_name': subject.class_field.department.department_name,
                'batch': subject.class_field.batch,
                'semester': subject.class_field.semester,
                'section': subject.class_field.section,
                'student_count': student_count
            })

        return Response(result)


# ============== REPORTS APIs ==============

class DashboardStatsView(APIView):
    """Get dashboard statistics for teacher"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.user

        # Get classes
        created_classes = ClassModel.objects.filter(created_by__teacher_email=teacher_email)
        taught_subjects = Subject.objects.filter(teacher__teacher_email=teacher_email)
        taught_class_ids = list(taught_subjects.values_list('class_field_id', flat=True))

        total_classes = ClassModel.objects.filter(
            Q(created_by__teacher_email=teacher_email) | Q(class_id__in=taught_class_ids)
        ).distinct().count()

        # Get students (from both class_field AND StudentClass enrollments)
        primary_students = Student.objects.filter(
            class_field_id__in=taught_class_ids
        ).values_list('student_email', flat=True)

        enrolled_students = StudentClass.objects.filter(
            class_obj_id__in=taught_class_ids
        ).values_list('student__student_email', flat=True)

        all_student_emails = set(primary_students) | set(enrolled_students)
        total_students = len(all_student_emails)

        # Total subjects
        total_subjects = taught_subjects.count()

        # Today's sessions
        today = timezone.now().date()
        todays_sessions = Session.objects.filter(
            period__subject__teacher__teacher_email=teacher_email,
            date=today
        ).count()

        # Average attendance
        all_attendance = Attendance.objects.filter(
            session__period__subject__teacher__teacher_email=teacher_email
        )
        total_records = all_attendance.count()
        present_records = all_attendance.filter(status='P').count()
        avg_attendance = (present_records / total_records * 100) if total_records > 0 else 0

        return Response({
            'total_classes': total_classes,
            'total_students': total_students,
            'total_subjects': total_subjects,
            'todays_sessions': todays_sessions,
            'average_attendance': round(avg_attendance, 2)
        })


class ClassReportView(APIView):
    """Get attendance report for a specific class"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get subjects for this class taught by this teacher
        subjects = Subject.objects.filter(
            class_field=class_obj,
            teacher__teacher_email=request.user
        ).select_related('course')

        subject_id = request.query_params.get('subject_id')
        if subject_id:
            subjects = subjects.filter(subject_id=subject_id)

        # Get students from both class_field and StudentClass
        primary_students = Student.objects.filter(class_field=class_obj)
        enrolled_student_emails = StudentClass.objects.filter(
            class_obj=class_obj
        ).values_list('student__student_email', flat=True)

        all_student_emails = set(primary_students.values_list('student_email', flat=True)) | set(enrolled_student_emails)
        students = Student.objects.filter(student_email__in=all_student_emails)

        # Get roll numbers from enrollments
        enrollment_roll_nos = {
            sc.student.student_email: sc.roll_no
            for sc in StudentClass.objects.filter(class_obj=class_obj).select_related('student')
        }

        reports = []
        for subject in subjects:
            sessions = Session.objects.filter(period__subject=subject)
            total_sessions = sessions.count()

            student_data = []
            total_attendance = 0

            for student in students:
                attendance = Attendance.objects.filter(
                    session__in=sessions,
                    student=student
                )
                present = attendance.filter(status='P').count()
                absent = attendance.filter(status='A').count()
                percentage = (present / total_sessions * 100) if total_sessions > 0 else 0
                total_attendance += percentage

                # Prefer roll_no from enrollment
                roll_no = enrollment_roll_nos.get(student.student_email, student.roll_no)

                student_data.append({
                    'student_email': student.student_email,
                    'student_name': student.name,
                    'roll_no': roll_no,
                    'total_sessions': total_sessions,
                    'present': present,
                    'absent': absent,
                    'percentage': round(percentage, 2)
                })

            avg_attendance = (total_attendance / len(students)) if students else 0

            reports.append({
                'class_id': class_obj.class_id,
                'class_name': str(class_obj),
                'subject_id': subject.subject_id,
                'subject_name': subject.course.course_name,
                'total_sessions': total_sessions,
                'average_attendance': round(avg_attendance, 2),
                'students': student_data
            })

        return Response(reports)


class SubjectReportView(APIView):
    """Get attendance report for a specific subject"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            subject = Subject.objects.get(subject_id=pk, teacher__teacher_email=request.user)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found or unauthorized'}, status=status.HTTP_404_NOT_FOUND)

        sessions = Session.objects.filter(period__subject=subject)
        total_sessions = sessions.count()

        # Get students from both class_field and StudentClass
        primary_students = Student.objects.filter(class_field=subject.class_field)
        enrolled_student_emails = StudentClass.objects.filter(
            class_obj=subject.class_field
        ).values_list('student__student_email', flat=True)

        all_student_emails = set(primary_students.values_list('student_email', flat=True)) | set(enrolled_student_emails)
        students = Student.objects.filter(student_email__in=all_student_emails)

        # Get roll numbers from enrollments
        enrollment_roll_nos = {
            sc.student.student_email: sc.roll_no
            for sc in StudentClass.objects.filter(class_obj=subject.class_field).select_related('student')
        }

        student_data = []
        total_attendance = 0

        for student in students:
            attendance = Attendance.objects.filter(
                session__in=sessions,
                student=student
            )
            present = attendance.filter(status='P').count()
            absent = attendance.filter(status='A').count()
            percentage = (present / total_sessions * 100) if total_sessions > 0 else 0
            total_attendance += percentage

            # Prefer roll_no from enrollment
            roll_no = enrollment_roll_nos.get(student.student_email, student.roll_no)

            student_data.append({
                'student_email': student.student_email,
                'student_name': student.name,
                'roll_no': roll_no,
                'total_sessions': total_sessions,
                'present': present,
                'absent': absent,
                'percentage': round(percentage, 2)
            })

        avg_attendance = (total_attendance / len(students)) if students else 0

        return Response({
            'subject_id': subject.subject_id,
            'subject_name': subject.course.course_name,
            'class_name': str(subject.class_field),
            'total_sessions': total_sessions,
            'average_attendance': round(avg_attendance, 2),
            'students': student_data
        })


# ============== PROFILE APIs ==============

class TeacherProfileView(APIView):
    """Get and update teacher profile"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            teacher = Teacher.objects.select_related('department', 'organization').get(teacher_email=request.user)
            data = TeacherProfileSerializer(teacher).data

            # Add organization info
            if teacher.organization:
                data['organization'] = {
                    'organization_id': teacher.organization.organization_id,
                    'name': teacher.organization.name,
                    'code': teacher.organization.code,
                }
            else:
                data['organization'] = None

            # Check if teacher is an admin
            is_admin = OrganizationAdmin.objects.filter(
                teacher__teacher_email=str(request.user)
            ).exists()
            data['is_admin'] = is_admin

            return Response(data)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            teacher = Teacher.objects.get(teacher_email=request.user)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

        # Handle department_name by looking up the department
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if 'department_name' in data and not data.get('department'):
            try:
                dept = Department.objects.get(department_name=data['department_name'])
                data['department'] = dept.department_id
            except Department.DoesNotExist:
                pass  # Ignore if department not found
            if 'department_name' in data:
                del data['department_name']

        serializer = TeacherProfileSerializer(teacher, data=data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer.save()
        return Response(serializer.data)


class StudentProfileView(APIView):
    """Get student profile"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Students only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.select_related(
                'class_field', 'class_field__department', 'class_field__department__organization'
            ).get(student_email=request.user)
            return Response(StudentProfileSerializer(student).data)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)


class StudentClassesView(APIView):
    """Get student's classes with attendance"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Students only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.get(student_email=request.user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get classes through enrollment + primary class
        enrolled_classes = StudentClass.objects.filter(student=student)
        enrolled_class_ids = list(enrolled_classes.values_list('class_obj_id', flat=True))

        # Create a map of class_id -> roll_no from enrollments
        roll_no_map = {e.class_obj_id: e.roll_no for e in enrolled_classes}

        class_ids = enrolled_class_ids + ([student.class_field_id] if student.class_field_id else [])
        classes = ClassModel.objects.filter(class_id__in=class_ids, is_active=True).select_related('department').distinct()

        result = []
        for class_obj in classes:
            # Get subjects in this class
            subjects = Subject.objects.filter(class_field=class_obj).select_related('course', 'teacher')

            subject_data = []
            total_present = 0
            total_absent = 0
            total_attended_sessions = 0  # Sessions where student has attendance record

            for subject in subjects:
                sessions = Session.objects.filter(period__subject=subject)

                # Get attendance records for this student in this subject
                attendance = Attendance.objects.filter(
                    session__in=sessions,
                    student=student
                )
                present = attendance.filter(status='P').count()
                absent = attendance.filter(status='A').count()
                attended_sessions = present + absent  # Sessions where student has a record

                total_present += present
                total_absent += absent
                total_attended_sessions += attended_sessions

                subject_data.append({
                    'subject_id': subject.subject_id,
                    'course_name': subject.course.course_name,
                    'teacher_name': subject.teacher.name,
                    'total_sessions': attended_sessions,  # Only sessions student attended
                    'present': present,
                    'absent': absent,
                    'percentage': round((present / attended_sessions * 100), 2) if attended_sessions > 0 else 0
                })

            overall_attendance = (total_present / total_attended_sessions * 100) if total_attended_sessions > 0 else 0

            # Get roll_no - prefer from enrollment, fallback to student's default
            roll_no = roll_no_map.get(class_obj.class_id, student.roll_no)

            result.append({
                'class_id': class_obj.class_id,
                'class_name': str(class_obj),
                'department_name': class_obj.department.department_name,
                'batch': class_obj.batch,
                'semester': class_obj.semester,
                'section': class_obj.section,
                'roll_no': roll_no,
                'subjects': subject_data,
                'overall_percentage': round(overall_attendance, 2),
                'total_present': total_present,
                'total_absent': total_absent,
                'total_sessions': total_attended_sessions
            })

        return Response(result)


class StudentTodayScheduleView(APIView):
    """Get today's schedule for a student"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Students only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.get(student_email=request.user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get student's classes
        enrolled_classes = StudentClass.objects.filter(student=student).values_list('class_obj_id', flat=True)
        class_ids = list(enrolled_classes) + ([student.class_field_id] if student.class_field_id else [])

        today = timezone.now().weekday()  # Monday = 0

        # Get periods for today
        periods = Period.objects.filter(
            subject__class_field__class_id__in=class_ids,
            subject__class_field__is_active=True,
            day_of_week=today
        ).select_related(
            'subject__course',
            'subject__teacher',
            'subject__class_field',
            'subject__class_field__department'
        ).order_by('period_no')

        schedule = []
        for period in periods:
            # Check if session exists for today
            today_date = timezone.now().date()
            session = Session.objects.filter(
                period=period,
                date=today_date
            ).first()

            attendance_status = None
            if session:
                att = Attendance.objects.filter(session=session, student=student).first()
                if att:
                    attendance_status = att.status

            schedule.append({
                'period_id': period.period_id,
                'period_no': period.period_no,
                'subject_id': period.subject.subject_id,
                'course_name': period.subject.course.course_name,
                'teacher_name': period.subject.teacher.name,
                'class_name': str(period.subject.class_field),
                'class_id': period.subject.class_field.class_id,
                'department_name': period.subject.class_field.department.department_name,
                'session_active': session.is_active if session else False,
                'session_id': session.session_id if session else None,
                'attendance_status': attendance_status
            })

        return Response(schedule)


class StudentWeeklyScheduleView(APIView):
    """Get full weekly schedule for a student"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Students only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.get(student_email=request.user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get student's classes
        enrolled_classes = StudentClass.objects.filter(student=student).values_list('class_obj_id', flat=True)
        class_ids = list(enrolled_classes) + ([student.class_field_id] if student.class_field_id else [])

        # Get all periods for the week (0=Monday to 5=Saturday)
        periods = Period.objects.filter(
            subject__class_field__class_id__in=class_ids,
            subject__class_field__is_active=True,
        ).select_related(
            'subject__course',
            'subject__teacher',
            'subject__class_field',
            'subject__class_field__department'
        ).order_by('day_of_week', 'period_no')

        schedule = []
        for period in periods:
            schedule.append({
                'period_id': period.period_id,
                'period_no': period.period_no,
                'day_of_week': period.day_of_week,
                'subject_id': period.subject.subject_id,
                'course_name': period.subject.course.course_name,
                'teacher_name': period.subject.teacher.name,
                'class_name': str(period.subject.class_field),
                'class_id': period.subject.class_field.class_id,
                'department_name': period.subject.class_field.department.department_name,
            })

        return Response(schedule)


class StudentDashboardStatsView(APIView):
    """Get dashboard statistics for student"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Students only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.get(student_email=request.user)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get student's classes
        enrolled_classes = StudentClass.objects.filter(student=student).values_list('class_obj_id', flat=True)
        class_ids = list(enrolled_classes) + ([student.class_field_id] if student.class_field_id else [])

        # Count active classes
        active_classes = ClassModel.objects.filter(
            class_id__in=class_ids,
            is_active=True
        ).count()

        # Count subjects
        total_subjects = Subject.objects.filter(
            class_field__class_id__in=class_ids,
            class_field__is_active=True
        ).count()

        # Today's classes count
        today = timezone.now().weekday()
        today_classes = Period.objects.filter(
            subject__class_field__class_id__in=class_ids,
            subject__class_field__is_active=True,
            day_of_week=today
        ).count()

        # Overall attendance - only count sessions where student has attendance record
        all_sessions = Session.objects.filter(
            period__subject__class_field__class_id__in=class_ids
        )

        # Get attendance records for this student
        student_attendance = Attendance.objects.filter(
            student=student,
            session__in=all_sessions
        )

        present_count = student_attendance.filter(status='P').count()
        absent_count = student_attendance.filter(status='A').count()

        # Total sessions = sessions where student actually has a record
        total_attended_sessions = present_count + absent_count

        overall_percentage = (present_count / total_attended_sessions * 100) if total_attended_sessions > 0 else 0

        return Response({
            'enrolled_classes': active_classes,
            'total_subjects': total_subjects,
            'today_classes': today_classes,
            'total_sessions': total_attended_sessions,
            'total_present': present_count,
            'total_absent': absent_count,
            'overall_percentage': round(overall_percentage, 2)
        })


class ExportClassAttendanceView(APIView):
    """Export attendance data for a class as JSON (can be converted to CSV/Excel on frontend)"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('coordinator').get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        teacher_email = str(request.user)
        is_coordinator = class_obj.coordinator and class_obj.coordinator.teacher_email == teacher_email

        # Get subjects - coordinators see all, teachers only see their own
        if is_coordinator:
            subjects = Subject.objects.filter(class_field=class_obj).select_related('course', 'teacher')
        else:
            subjects = Subject.objects.filter(
                class_field=class_obj,
                teacher__teacher_email=teacher_email
            ).select_related('course', 'teacher')

        if not subjects.exists():
            return Response({'error': 'No subjects found for export'}, status=status.HTTP_404_NOT_FOUND)

        # Get all students in this class
        enrollments = StudentClass.objects.filter(class_obj=class_obj).select_related('student')

        # Get class summary
        class_info = {
            'class_name': str(class_obj),
            'department': class_obj.department.department_name,
            'batch': class_obj.batch,
            'semester': class_obj.semester,
            'section': class_obj.section,
            'total_students': enrollments.count(),
            'total_subjects': subjects.count(),
            'is_active': class_obj.is_active,
            'exported_at': timezone.now().isoformat(),
        }

        # Get subject-wise session counts
        subject_data = []
        for subject in subjects:
            sessions = Session.objects.filter(period__subject=subject)
            subject_data.append({
                'subject_id': subject.subject_id,
                'course_name': subject.course.course_name,
                'teacher_name': subject.teacher.name,
                'total_sessions': sessions.count(),
            })

        # Get student-wise attendance
        students_data = []
        for enrollment in enrollments:
            student = enrollment.student
            student_attendance = {
                'roll_no': enrollment.roll_no,
                'name': student.name,
                'email': student.student_email,
                'subjects': []
            }

            total_present = 0
            total_absent = 0
            total_sessions = 0

            for subject in subjects:
                sessions = Session.objects.filter(period__subject=subject)
                session_count = sessions.count()

                if session_count > 0:
                    attendance = Attendance.objects.filter(
                        session__in=sessions,
                        student=student
                    )
                    present = attendance.filter(status='P').count()
                    absent = attendance.filter(status='A').count()

                    total_present += present
                    total_absent += absent
                    total_sessions += session_count

                    student_attendance['subjects'].append({
                        'course_name': subject.course.course_name,
                        'total_sessions': session_count,
                        'present': present,
                        'absent': absent,
                        'percentage': round((present / session_count * 100), 2) if session_count > 0 else 0
                    })

            student_attendance['total_sessions'] = total_sessions
            student_attendance['total_present'] = total_present
            student_attendance['total_absent'] = total_absent
            student_attendance['overall_percentage'] = round(
                (total_present / total_sessions * 100), 2
            ) if total_sessions > 0 else 0

            students_data.append(student_attendance)

        # Sort by roll number
        students_data.sort(key=lambda x: x['roll_no'])

        return Response({
            'class_info': class_info,
            'subjects': subject_data,
            'students': students_data
        })


class ClassDetailedStatsView(APIView):
    """Get detailed class statistics for display"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('coordinator').get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        teacher_email = str(request.user)
        is_coordinator = class_obj.coordinator and class_obj.coordinator.teacher_email == teacher_email

        # Get all subjects
        subjects = Subject.objects.filter(class_field=class_obj).select_related('course', 'teacher')

        # Get all sessions
        all_sessions = Session.objects.filter(period__subject__class_field=class_obj)
        total_sessions = all_sessions.count()

        # Get unique dates with sessions
        unique_dates = all_sessions.values_list('date', flat=True).distinct().count()

        # Get total students (from both class_field AND StudentClass enrollments)
        primary_students = Student.objects.filter(class_field=class_obj).values_list('student_email', flat=True)
        enrolled_students = StudentClass.objects.filter(class_obj=class_obj).values_list('student__student_email', flat=True)
        all_student_emails = set(primary_students) | set(enrolled_students)
        total_student_count = len(all_student_emails)

        # Calculate average attendance
        if total_sessions > 0 and total_student_count > 0:
            total_possible = total_sessions * total_student_count
            total_present = Attendance.objects.filter(
                session__in=all_sessions,
                status='P'
            ).count()
            avg_attendance = (total_present / total_possible * 100) if total_possible > 0 else 0
        else:
            avg_attendance = 0

        # Subject-wise stats
        subject_stats = []
        for subject in subjects:
            sessions = Session.objects.filter(period__subject=subject)
            session_count = sessions.count()

            if session_count > 0:
                present = Attendance.objects.filter(session__in=sessions, status='P').count()
                possible = session_count * total_student_count
                avg = (present / possible * 100) if possible > 0 else 0
            else:
                avg = 0

            is_subject_owner = subject.teacher.teacher_email == teacher_email
            can_export = is_subject_owner or is_coordinator

            subject_stats.append({
                'subject_id': subject.subject_id,
                'course_name': subject.course.course_name,
                'teacher_name': subject.teacher.name,
                'total_sessions': session_count,
                'average_attendance': round(avg, 2),
                'is_subject_owner': is_subject_owner,
                'can_export': can_export,
            })

        return Response({
            'class_id': class_obj.class_id,
            'class_name': str(class_obj),
            'department_name': class_obj.department.department_name,
            'batch': class_obj.batch,
            'semester': class_obj.semester,
            'section': class_obj.section,
            'is_active': class_obj.is_active,
            'total_students': total_student_count,
            'total_subjects': subjects.count(),
            'total_sessions': total_sessions,
            'classes_taken': unique_dates,
            'average_attendance': round(avg_attendance, 2),
            'subjects': subject_stats
        })


# ========== ANNOUNCEMENTS & NOTES API ==========

class AnnouncementListView(APIView):
    """List and create announcements for a class (Teacher only)"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can access this'}, status=403)

        class_id = request.query_params.get('class_id')
        if not class_id:
            # Return all announcements by this teacher
            announcements = Announcement.objects.filter(teacher__teacher_email=str(request.user))
        else:
            announcements = Announcement.objects.filter(
                class_obj_id=class_id,
                teacher__teacher_email=str(request.user)
            )

        data = []
        for ann in announcements:
            data.append({
                'announcement_id': ann.announcement_id,
                'class_id': ann.class_obj_id,
                'class_name': str(ann.class_obj),
                'subject_id': ann.subject_id,
                'subject_name': ann.subject.course.course_name if ann.subject else None,
                'title': ann.title,
                'content': ann.content,
                'created_at': ann.created_at.isoformat(),
                'updated_at': ann.updated_at.isoformat(),
            })

        return Response(data)

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can create announcements'}, status=403)

        class_id = request.data.get('class_id')
        subject_id = request.data.get('subject_id')
        title = request.data.get('title')
        content = request.data.get('content')

        if not class_id or not title or not content:
            return Response({'error': 'class_id, title and content are required'}, status=400)

        try:
            teacher = Teacher.objects.get(teacher_email=str(request.user))
            class_obj = ClassModel.objects.get(class_id=class_id)
            subject = Subject.objects.get(subject_id=subject_id) if subject_id else None

            announcement = Announcement.objects.create(
                class_obj=class_obj,
                subject=subject,
                teacher=teacher,
                title=title,
                content=content
            )

            return Response({
                'announcement_id': announcement.announcement_id,
                'message': 'Announcement created successfully'
            }, status=201)

        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=404)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=404)


class AnnouncementDetailView(APIView):
    """Update and delete a specific announcement"""
    permission_classes = [IsJWTAuthenticated]

    def put(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can update announcements'}, status=403)

        try:
            announcement = Announcement.objects.get(
                announcement_id=pk,
                teacher__teacher_email=str(request.user)
            )

            title = request.data.get('title')
            content = request.data.get('content')

            if title:
                announcement.title = title
            if content:
                announcement.content = content

            announcement.save()

            return Response({'message': 'Announcement updated'})

        except Announcement.DoesNotExist:
            return Response({'error': 'Announcement not found'}, status=404)

    def delete(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can delete announcements'}, status=403)

        try:
            announcement = Announcement.objects.get(
                announcement_id=pk,
                teacher__teacher_email=str(request.user)
            )
            announcement.delete()
            return Response({'message': 'Announcement deleted'})

        except Announcement.DoesNotExist:
            return Response({'error': 'Announcement not found'}, status=404)


class StudentAnnouncementsView(APIView):
    """Get announcements for student's enrolled classes"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'student':
            return Response({'error': 'Only students can access this'}, status=403)

        try:
            student = Student.objects.get(student_email=str(request.user))

            # Get all enrolled class IDs
            enrolled_class_ids = set()
            enrolled_class_ids.add(student.class_field_id)
            for enrollment in StudentClass.objects.filter(student=student):
                enrolled_class_ids.add(enrollment.class_obj_id)

            # Get announcements for these classes
            announcements = Announcement.objects.filter(class_obj_id__in=enrolled_class_ids)

            # Get read status
            read_ids = set(
                AnnouncementRead.objects.filter(student=student).values_list('announcement_id', flat=True)
            )

            data = []
            for ann in announcements:
                data.append({
                    'announcement_id': ann.announcement_id,
                    'class_id': ann.class_obj_id,
                    'class_name': str(ann.class_obj),
                    'subject_id': ann.subject_id,
                    'subject_name': ann.subject.course.course_name if ann.subject else None,
                    'teacher_name': ann.teacher.name,
                    'title': ann.title,
                    'content': ann.content,
                    'created_at': ann.created_at.isoformat(),
                    'is_read': ann.announcement_id in read_ids,
                })

            # Count unread
            unread_count = len([a for a in data if not a['is_read']])

            return Response({
                'announcements': data,
                'unread_count': unread_count
            })

        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=404)

    def post(self, request):
        """Mark announcement as read"""
        if request.user_type != 'student':
            return Response({'error': 'Only students can access this'}, status=403)

        announcement_id = request.data.get('announcement_id')
        if not announcement_id:
            return Response({'error': 'announcement_id is required'}, status=400)

        try:
            student = Student.objects.get(student_email=str(request.user))
            announcement = Announcement.objects.get(announcement_id=announcement_id)

            AnnouncementRead.objects.get_or_create(
                announcement=announcement,
                student=student
            )

            return Response({'message': 'Marked as read'})

        except (Student.DoesNotExist, Announcement.DoesNotExist):
            return Response({'error': 'Not found'}, status=404)


class TeacherNoteListView(APIView):
    """Private notes for teacher"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can access notes'}, status=403)

        class_id = request.query_params.get('class_id')

        notes = TeacherNote.objects.filter(teacher__teacher_email=str(request.user))
        if class_id:
            notes = notes.filter(class_obj_id=class_id)

        data = []
        for note in notes:
            data.append({
                'note_id': note.note_id,
                'class_id': note.class_obj_id,
                'class_name': str(note.class_obj),
                'subject_id': note.subject_id,
                'subject_name': note.subject.course.course_name if note.subject else None,
                'content': note.content,
                'created_at': note.created_at.isoformat(),
                'updated_at': note.updated_at.isoformat(),
            })

        return Response(data)

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can create notes'}, status=403)

        class_id = request.data.get('class_id')
        subject_id = request.data.get('subject_id')
        content = request.data.get('content')

        if not class_id or not content:
            return Response({'error': 'class_id and content are required'}, status=400)

        try:
            teacher = Teacher.objects.get(teacher_email=str(request.user))
            class_obj = ClassModel.objects.get(class_id=class_id)
            subject = Subject.objects.get(subject_id=subject_id) if subject_id else None

            note = TeacherNote.objects.create(
                class_obj=class_obj,
                subject=subject,
                teacher=teacher,
                content=content
            )

            return Response({
                'note_id': note.note_id,
                'message': 'Note created successfully'
            }, status=201)

        except (Teacher.DoesNotExist, ClassModel.DoesNotExist):
            return Response({'error': 'Not found'}, status=404)


class TeacherNoteDetailView(APIView):
    """Update/delete a specific note"""
    permission_classes = [IsJWTAuthenticated]

    def put(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can update notes'}, status=403)

        try:
            note = TeacherNote.objects.get(
                note_id=pk,
                teacher__teacher_email=str(request.user)
            )
            content = request.data.get('content')
            if content:
                note.content = content
                note.save()
            return Response({'message': 'Note updated'})

        except TeacherNote.DoesNotExist:
            return Response({'error': 'Note not found'}, status=404)

    def delete(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can delete notes'}, status=403)

        try:
            note = TeacherNote.objects.get(
                note_id=pk,
                teacher__teacher_email=str(request.user)
            )
            note.delete()
            return Response({'message': 'Note deleted'})

        except TeacherNote.DoesNotExist:
            return Response({'error': 'Note not found'}, status=404)


# ========== PERIOD TIME SLOTS API ==========

class PeriodTimeSlotListView(APIView):
    """Manage default time slots for periods"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can access time slots'}, status=403)

        slots = PeriodTimeSlot.objects.filter(teacher__teacher_email=str(request.user))

        data = []
        for slot in slots:
            data.append({
                'slot_id': slot.slot_id,
                'period_no': slot.period_no,
                'start_time': slot.start_time.strftime('%H:%M') if slot.start_time else None,
                'end_time': slot.end_time.strftime('%H:%M') if slot.end_time else None,
            })

        return Response(data)

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can create time slots'}, status=403)

        try:
            from datetime import datetime
            teacher = Teacher.objects.get(teacher_email=str(request.user))

            # Handle array of slots (from frontend)
            slots_data = request.data.get('slots', [])
            if not slots_data:
                # Fallback to single slot format
                period_no = request.data.get('period_no')
                start_time = request.data.get('start_time')
                end_time = request.data.get('end_time')
                if period_no and start_time and end_time:
                    slots_data = [{'period_no': period_no, 'start_time': start_time, 'end_time': end_time}]

            if not slots_data:
                return Response({'error': 'No slots provided'}, status=400)

            saved_slots = []
            for slot_data in slots_data:
                period_no = slot_data.get('period_no')
                start_time = slot_data.get('start_time')
                end_time = slot_data.get('end_time')

                if not period_no or not start_time or not end_time:
                    continue

                start = datetime.strptime(start_time, '%H:%M').time()
                end = datetime.strptime(end_time, '%H:%M').time()

                slot, created = PeriodTimeSlot.objects.update_or_create(
                    teacher=teacher,
                    period_no=period_no,
                    defaults={'start_time': start, 'end_time': end}
                )
                saved_slots.append({
                    'slot_id': slot.slot_id,
                    'period_no': slot.period_no,
                    'start_time': slot.start_time.strftime('%H:%M'),
                    'end_time': slot.end_time.strftime('%H:%M'),
                })

            return Response(saved_slots, status=200)

        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=404)
        except ValueError:
            return Response({'error': 'Invalid time format. Use HH:MM'}, status=400)

    def delete(self, request):
        """Delete time slot by period_no"""
        if request.user_type != 'teacher':
            return Response({'error': 'Only teachers can delete time slots'}, status=403)

        period_no = request.query_params.get('period_no')
        if not period_no:
            return Response({'error': 'period_no is required'}, status=400)

        deleted, _ = PeriodTimeSlot.objects.filter(
            teacher__teacher_email=str(request.user),
            period_no=period_no
        ).delete()

        if deleted:
            return Response({'message': 'Time slot deleted'})
        return Response({'error': 'Time slot not found'}, status=404)


# ============== COORDINATOR / CLASS TEACHER MANAGEMENT ==============

class ClassTeacherListView(APIView):
    """List and manage teachers assigned to a class"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        """Get all teachers assigned to a class"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.select_related('coordinator').get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get assigned teachers
        assigned = ClassTeacher.objects.filter(class_obj=class_obj).select_related('teacher', 'added_by')

        teachers = []
        for assignment in assigned:
            teachers.append({
                'teacher_email': assignment.teacher.teacher_email,
                'teacher_name': assignment.teacher.name,
                'designation': assignment.teacher.designation,
                'department_name': assignment.teacher.department.department_name,
                'added_by': assignment.added_by.name if assignment.added_by else None,
                'added_at': assignment.added_at.isoformat(),
            })

        # Add coordinator if not in assigned teachers
        if class_obj.coordinator:
            coordinator_emails = [t['teacher_email'] for t in teachers]
            if class_obj.coordinator.teacher_email not in coordinator_emails:
                teachers.insert(0, {
                    'teacher_email': class_obj.coordinator.teacher_email,
                    'teacher_name': class_obj.coordinator.name,
                    'designation': class_obj.coordinator.designation,
                    'department_name': class_obj.coordinator.department.department_name,
                    'added_by': 'Coordinator',
                    'added_at': class_obj.created_at.isoformat() if class_obj.created_at else None,
                    'is_coordinator': True,
                })

        return Response({
            'class_id': class_obj.class_id,
            'class_name': str(class_obj),
            'coordinator_email': class_obj.coordinator.teacher_email if class_obj.coordinator else None,
            'coordinator_name': class_obj.coordinator.name if class_obj.coordinator else None,
            'teachers': teachers
        })

    def post(self, request, pk):
        """Add a teacher to a class (coordinator only)"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if user is coordinator
        if not class_obj.coordinator or class_obj.coordinator.teacher_email != request.user:
            return Response({'error': 'Only the class coordinator can add teachers'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.data.get('teacher_email')
        if not teacher_email:
            return Response({'error': 'teacher_email is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            teacher = Teacher.objects.get(teacher_email=teacher_email)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

        # Auto-verify teacher when added by coordinator
        if not teacher.verified:
            teacher.verified = True
            teacher.save()

        # Check if already assigned
        if ClassTeacher.objects.filter(class_obj=class_obj, teacher=teacher).exists():
            return Response({'error': 'Teacher already assigned to this class'}, status=status.HTTP_400_BAD_REQUEST)

        coordinator = Teacher.objects.get(teacher_email=request.user)
        assignment = ClassTeacher.objects.create(
            class_obj=class_obj,
            teacher=teacher,
            added_by=coordinator
        )

        return Response({
            'message': 'Teacher added and verified successfully',
            'teacher_email': teacher.teacher_email,
            'teacher_name': teacher.name,
            'verified': teacher.verified,
            'added_at': assignment.added_at.isoformat()
        }, status=status.HTTP_201_CREATED)

    def delete(self, request, pk):
        """Remove a teacher from a class (coordinator only)"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            class_obj = ClassModel.objects.get(class_id=pk)
        except ClassModel.DoesNotExist:
            return Response({'error': 'Class not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if user is coordinator
        if not class_obj.coordinator or class_obj.coordinator.teacher_email != request.user:
            return Response({'error': 'Only the class coordinator can remove teachers'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = request.query_params.get('teacher_email')
        if not teacher_email:
            return Response({'error': 'teacher_email is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Can't remove the coordinator
        if teacher_email == class_obj.coordinator.teacher_email:
            return Response({'error': 'Cannot remove the coordinator from the class'}, status=status.HTTP_400_BAD_REQUEST)

        deleted, _ = ClassTeacher.objects.filter(
            class_obj=class_obj,
            teacher__teacher_email=teacher_email
        ).delete()

        if deleted:
            return Response({'message': 'Teacher removed from class'})
        return Response({'error': 'Teacher not found in this class'}, status=status.HTTP_404_NOT_FOUND)


class TeacherSearchView(APIView):
    """Search for teachers by email or name"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        query = request.query_params.get('q', '').strip()
        if len(query) < 2:
            return Response({'error': 'Query must be at least 2 characters'}, status=status.HTTP_400_BAD_REQUEST)

        # Search all teachers (including unverified) so coordinators can find and add them
        teachers = Teacher.objects.filter(
            Q(teacher_email__icontains=query) | Q(name__icontains=query)
        ).select_related('department')[:20]

        return Response([{
            'teacher_email': t.teacher_email,
            'name': t.name,
            'designation': t.designation,
            'department_name': t.department.department_name,
            'verified': t.verified,
        } for t in teachers])


class CreateTeacherView(APIView):
    """Create a new teacher (for coordinators)"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        email = request.data.get('email', '').strip().lower()
        name = request.data.get('name', '').strip()
        designation = request.data.get('designation', 'Faculty').strip()

        if not email or not name:
            return Response({'error': 'Email and name are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if teacher already exists
        if Teacher.objects.filter(teacher_email=email).exists():
            return Response({'error': 'Teacher with this email already exists'}, status=status.HTTP_400_BAD_REQUEST)

        # Get the creating teacher's info
        try:
            creating_teacher = Teacher.objects.get(teacher_email=str(request.user))
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

        # Create new teacher with same organization and department
        # Auto-verify since coordinator is vouching for them
        teacher = Teacher.objects.create(
            teacher_email=email,
            name=name,
            designation=designation,
            organization=creating_teacher.organization,
            department=creating_teacher.department,
            verified=True  # Auto-verified when created by coordinator
        )

        return Response({
            'teacher_email': teacher.teacher_email,
            'name': teacher.name,
            'designation': teacher.designation,
            'department_name': teacher.department.department_name,
            'verified': teacher.verified,
            'message': 'Teacher created and verified. They can now login with Google.'
        }, status=status.HTTP_201_CREATED)


class SubjectEnrollmentView(APIView):
    """Manage student enrollment in subjects"""
    permission_classes = [IsJWTAuthenticated]

    def _can_manage_subject(self, subject, teacher_email):
        """Check if teacher can manage this subject (is owner or coordinator)"""
        is_subject_owner = subject.teacher.teacher_email == teacher_email
        is_coordinator = subject.class_field.coordinator and subject.class_field.coordinator.teacher_email == teacher_email
        return is_subject_owner or is_coordinator

    def get(self, request, pk):
        """Get students enrolled in a subject"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            subject = Subject.objects.select_related('teacher', 'class_field__coordinator').get(subject_id=pk)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Get enrolled students
        enrollments = SubjectEnrollment.objects.filter(subject=subject).select_related('student')

        # Also get class students who are NOT enrolled in this subject
        class_students = StudentClass.objects.filter(class_obj=subject.class_field).select_related('student')
        enrolled_emails = set(e.student.student_email for e in enrollments)

        not_enrolled = [
            {
                'student_email': sc.student.student_email,
                'name': sc.student.name,
                'roll_no': sc.roll_no,
                'verified': sc.student.verified,
            }
            for sc in class_students if sc.student.student_email not in enrolled_emails
        ]

        return Response({
            'subject_id': subject.subject_id,
            'course_name': subject.course.course_name,
            'class_name': str(subject.class_field),
            'enrolled': [{
                'student_email': e.student.student_email,
                'name': e.student.name,
                'roll_no': e.student.roll_no if hasattr(e.student, 'roll_no') else '',
                'verified': e.student.verified,
                'enrolled_at': e.enrolled_at.isoformat(),
            } for e in enrollments],
            'not_enrolled': not_enrolled,
        })

    def post(self, request, pk):
        """Enroll a student in a subject"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            subject = Subject.objects.select_related('teacher', 'class_field__coordinator').get(subject_id=pk)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if teacher can manage this subject (is owner or coordinator)
        if not self._can_manage_subject(subject, str(request.user)):
            return Response(
                {'error': 'Unauthorized - only subject teacher or class coordinator can enroll students'},
                status=status.HTTP_403_FORBIDDEN
            )

        student_email = request.data.get('student_email')
        if not student_email:
            return Response({'error': 'Student email is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student = Student.objects.get(student_email=student_email)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if already enrolled
        if SubjectEnrollment.objects.filter(subject=subject, student=student).exists():
            return Response({'error': 'Student already enrolled'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            teacher = Teacher.objects.get(teacher_email=str(request.user))
        except Teacher.DoesNotExist:
            teacher = None

        enrollment = SubjectEnrollment.objects.create(
            subject=subject,
            student=student,
            enrolled_by=teacher
        )

        return Response({
            'message': 'Student enrolled successfully',
            'enrollment_id': enrollment.enrollment_id,
        }, status=status.HTTP_201_CREATED)

    def delete(self, request, pk):
        """Remove a student from a subject"""
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            subject = Subject.objects.select_related('teacher', 'class_field__coordinator').get(subject_id=pk)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if teacher can manage this subject (is owner or coordinator)
        if not self._can_manage_subject(subject, str(request.user)):
            return Response(
                {'error': 'Unauthorized - only subject teacher or class coordinator can remove students'},
                status=status.HTTP_403_FORBIDDEN
            )

        student_email = request.query_params.get('student_email')
        if not student_email:
            return Response({'error': 'Student email is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            enrollment = SubjectEnrollment.objects.get(subject_id=pk, student__student_email=student_email)
            enrollment.delete()
            return Response({'message': 'Student removed from subject'})
        except SubjectEnrollment.DoesNotExist:
            return Response({'error': 'Enrollment not found'}, status=status.HTTP_404_NOT_FOUND)


class EnrollAllStudentsView(APIView):
    """Enroll all class students in a subject at once"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        try:
            subject = Subject.objects.select_related('teacher', 'class_field__coordinator').get(subject_id=pk)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if teacher can manage this subject (is owner or coordinator)
        teacher_email = str(request.user)
        is_subject_owner = subject.teacher.teacher_email == teacher_email
        is_coordinator = subject.class_field.coordinator and subject.class_field.coordinator.teacher_email == teacher_email

        if not (is_subject_owner or is_coordinator):
            return Response(
                {'error': 'Unauthorized - only subject teacher or class coordinator can enroll students'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            teacher = Teacher.objects.get(teacher_email=teacher_email)
        except Teacher.DoesNotExist:
            teacher = None

        # Get all students in the class
        class_students = StudentClass.objects.filter(class_obj=subject.class_field).select_related('student')

        # Get already enrolled
        already_enrolled = set(
            SubjectEnrollment.objects.filter(subject=subject).values_list('student__student_email', flat=True)
        )

        # Enroll students who aren't already enrolled
        enrolled_count = 0
        for sc in class_students:
            if sc.student.student_email not in already_enrolled:
                SubjectEnrollment.objects.create(
                    subject=subject,
                    student=sc.student,
                    enrolled_by=teacher
                )
                enrolled_count += 1

        return Response({
            'message': f'{enrolled_count} students enrolled',
            'enrolled_count': enrolled_count,
        })


# ============== ORGANIZATION ADMIN APIs ==============

class AdminCheckView(APIView):
    """Check if current user is an organization admin"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'is_admin': False, 'organization': None})

        try:
            admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
                teacher__teacher_email=str(request.user)
            ).first()

            if admin_entry:
                return Response({
                    'is_admin': True,
                    'organization': {
                        'organization_id': admin_entry.organization.organization_id,
                        'name': admin_entry.organization.name,
                        'code': admin_entry.organization.code,
                    }
                })
            else:
                return Response({'is_admin': False, 'organization': None})
        except Exception:
            return Response({'is_admin': False, 'organization': None})


class AdminDashboardView(APIView):
    """Get organization dashboard stats for admin"""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        # Check if user is admin
        admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
            teacher__teacher_email=str(request.user)
        ).first()

        if not admin_entry:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        org = admin_entry.organization

        # Get stats
        teacher_count = Teacher.objects.filter(organization=org).count()
        department_count = Department.objects.filter(organization=org).count()
        class_count = ClassModel.objects.filter(department__organization=org, is_active=True).count()
        student_count = Student.objects.filter(class_field__department__organization=org).distinct().count()

        # Get admin list
        admins = OrganizationAdmin.objects.filter(organization=org).select_related('teacher')
        admin_list = [
            {
                'email': a.teacher.teacher_email,
                'name': a.teacher.name,
                'added_at': a.added_at.isoformat() if a.added_at else None,
            }
            for a in admins
        ]

        return Response({
            'organization': {
                'organization_id': org.organization_id,
                'name': org.name,
                'code': org.code,
                'address': org.address,
            },
            'stats': {
                'teachers': teacher_count,
                'departments': department_count,
                'active_classes': class_count,
                'students': student_count,
            },
            'admins': admin_list,
        })


class AdminTeacherListView(APIView):
    """List, add, and delete teachers in organization"""
    permission_classes = [IsJWTAuthenticated]

    def _get_org_for_admin(self, request):
        """Get organization if user is admin"""
        if request.user_type != 'teacher':
            return None
        admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
            teacher__teacher_email=str(request.user)
        ).first()
        return admin_entry.organization if admin_entry else None

    def get(self, request):
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        teachers = Teacher.objects.filter(organization=org).select_related('department')

        teacher_list = []
        for t in teachers:
            # Check if this teacher is also an admin
            is_admin = OrganizationAdmin.objects.filter(organization=org, teacher=t).exists()
            teacher_list.append({
                'email': t.teacher_email,
                'name': t.name,
                'designation': t.designation,
                'department_id': t.department.department_id if t.department else None,
                'department_name': t.department.department_name if t.department else None,
                'verified': t.verified,
                'is_admin': is_admin,
            })

        return Response({
            'organization_id': org.organization_id,
            'organization_name': org.name,
            'teachers': teacher_list,
            'count': len(teacher_list),
        })

    def post(self, request):
        """Add a new teacher to the organization"""
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        email = request.data.get('email')
        name = request.data.get('name')
        designation = request.data.get('designation', 'Professor')
        department_id = request.data.get('department_id')
        make_admin = request.data.get('make_admin', False)

        if not email or not name:
            return Response({'error': 'Email and name are required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if teacher already exists
        if Teacher.objects.filter(teacher_email=email).exists():
            return Response({'error': 'Teacher with this email already exists'}, status=status.HTTP_400_BAD_REQUEST)

        # Get department
        try:
            department = Department.objects.get(department_id=department_id, organization=org)
        except Department.DoesNotExist:
            return Response({'error': 'Department not found in this organization'}, status=status.HTTP_400_BAD_REQUEST)

        # Create teacher
        teacher = Teacher.objects.create(
            teacher_email=email,
            name=name,
            designation=designation,
            department=department,
            organization=org,
            verified=True,  # Admin-added teachers are pre-verified
        )

        # Make admin if requested
        if make_admin:
            OrganizationAdmin.objects.create(
                organization=org,
                teacher=teacher,
                added_by=Teacher.objects.filter(teacher_email=str(request.user)).first()
            )

        return Response({
            'message': 'Teacher added successfully',
            'email': teacher.teacher_email,
            'name': teacher.name,
            'is_admin': make_admin,
        }, status=status.HTTP_201_CREATED)

    def delete(self, request):
        """Remove a teacher from the organization"""
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        email = request.query_params.get('email')
        if not email:
            return Response({'error': 'Teacher email is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Prevent self-deletion
        if email == str(request.user):
            return Response({'error': 'Cannot delete your own account'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            teacher = Teacher.objects.get(teacher_email=email, organization=org)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found in this organization'}, status=status.HTTP_404_NOT_FOUND)

        # Remove admin status first if exists
        OrganizationAdmin.objects.filter(organization=org, teacher=teacher).delete()

        # Delete teacher
        teacher.delete()

        return Response({'message': 'Teacher removed successfully'})


class AdminTeacherImportView(APIView):
    """Import teachers from CSV/Excel file"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
            teacher__teacher_email=str(request.user)
        ).first()

        if not admin_entry:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        org = admin_entry.organization

        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        uploaded_file = request.FILES['file']
        filename = uploaded_file.name.lower()

        try:
            if filename.endswith('.csv'):
                # Handle CSV
                decoded_file = uploaded_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                # Handle Excel
                try:
                    import openpyxl
                except ImportError:
                    return Response({'error': 'Excel support not available'}, status=status.HTTP_400_BAD_REQUEST)

                wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        rows.append({headers[i]: (row[i] if i < len(row) else '') for i in range(len(headers))})
            else:
                return Response({'error': 'Invalid file type. Use CSV or Excel'}, status=status.HTTP_400_BAD_REQUEST)

            added = 0
            skipped = 0
            errors = []

            for row in rows:
                email = row.get('email', '').strip()
                name = row.get('name', '').strip()
                designation = row.get('designation', 'Professor').strip()
                department_name = row.get('department', '').strip()

                if not email or not name:
                    skipped += 1
                    continue

                # Check if teacher exists
                if Teacher.objects.filter(teacher_email=email).exists():
                    skipped += 1
                    errors.append(f'{email}: already exists')
                    continue

                # Find or create department
                department = None
                if department_name:
                    department = Department.objects.filter(
                        organization=org,
                        department_name__iexact=department_name
                    ).first()
                    if not department:
                        # Create department if doesn't exist
                        department = Department.objects.create(
                            organization=org,
                            department_name=department_name
                        )

                if not department:
                    # Use first department in org as fallback
                    department = Department.objects.filter(organization=org).first()
                    if not department:
                        skipped += 1
                        errors.append(f'{email}: no department found')
                        continue

                # Create teacher
                Teacher.objects.create(
                    teacher_email=email,
                    name=name,
                    designation=designation,
                    department=department,
                    organization=org,
                    verified=True,
                )
                added += 1

            return Response({
                'message': f'{added} teachers imported, {skipped} skipped',
                'added': added,
                'skipped': skipped,
                'errors': errors[:10],  # Return first 10 errors
            })

        except Exception as e:
            return Response({'error': f'Failed to parse file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class AdminDepartmentListView(APIView):
    """List, add, and delete departments in organization"""
    permission_classes = [IsJWTAuthenticated]

    def _get_org_for_admin(self, request):
        """Get organization if user is admin"""
        if request.user_type != 'teacher':
            return None
        admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
            teacher__teacher_email=str(request.user)
        ).first()
        return admin_entry.organization if admin_entry else None

    def get(self, request):
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        departments = Department.objects.filter(organization=org).annotate(
            teacher_count=Count('teacher'),
            class_count=Count('class', filter=Q(class__is_active=True))
        )

        dept_list = [
            {
                'department_id': d.department_id,
                'department_name': d.department_name,
                'school': d.school,
                'teacher_count': d.teacher_count,
                'class_count': d.class_count,
            }
            for d in departments
        ]

        return Response({
            'organization_id': org.organization_id,
            'organization_name': org.name,
            'departments': dept_list,
            'count': len(dept_list),
        })

    def post(self, request):
        """Add a new department"""
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        name = request.data.get('department_name')
        school = request.data.get('school', '')

        if not name:
            return Response({'error': 'Department name is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Check if department exists
        if Department.objects.filter(organization=org, department_name__iexact=name).exists():
            return Response({'error': 'Department already exists'}, status=status.HTTP_400_BAD_REQUEST)

        department = Department.objects.create(
            organization=org,
            department_name=name,
            school=school,
        )

        return Response({
            'message': 'Department created',
            'department_id': department.department_id,
            'department_name': department.department_name,
        }, status=status.HTTP_201_CREATED)

    def delete(self, request):
        """Delete a department"""
        org = self._get_org_for_admin(request)
        if not org:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        dept_id = request.query_params.get('department_id')
        if not dept_id:
            return Response({'error': 'Department ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            department = Department.objects.get(department_id=dept_id, organization=org)
        except Department.DoesNotExist:
            return Response({'error': 'Department not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if department has teachers or classes
        if Teacher.objects.filter(department=department).exists():
            return Response(
                {'error': 'Cannot delete department with teachers. Reassign teachers first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if ClassModel.objects.filter(department=department).exists():
            return Response(
                {'error': 'Cannot delete department with classes. Delete classes first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        department.delete()
        return Response({'message': 'Department deleted'})


class AdminToggleView(APIView):
    """Toggle admin status for a teacher"""
    permission_classes = [IsJWTAuthenticated]

    def post(self, request):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        admin_entry = OrganizationAdmin.objects.select_related('organization').filter(
            teacher__teacher_email=str(request.user)
        ).first()

        if not admin_entry:
            return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

        org = admin_entry.organization
        teacher_email = request.data.get('teacher_email')

        if not teacher_email:
            return Response({'error': 'Teacher email is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Prevent self-demotion
        if teacher_email == str(request.user):
            return Response({'error': 'Cannot change your own admin status'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            teacher = Teacher.objects.get(teacher_email=teacher_email, organization=org)
        except Teacher.DoesNotExist:
            return Response({'error': 'Teacher not found in organization'}, status=status.HTTP_404_NOT_FOUND)

        # Toggle admin status
        existing = OrganizationAdmin.objects.filter(organization=org, teacher=teacher).first()
        if existing:
            existing.delete()
            return Response({'message': 'Admin status removed', 'is_admin': False})
        else:
            OrganizationAdmin.objects.create(
                organization=org,
                teacher=teacher,
                added_by=Teacher.objects.filter(teacher_email=str(request.user)).first()
            )
            return Response({'message': 'Admin status granted', 'is_admin': True})


class ExportSubjectAttendanceView(APIView):
    """Export attendance register for a subject — supports months listing, monthly date-wise, and semester summary."""
    permission_classes = [IsJWTAuthenticated]

    def get(self, request, pk):
        if request.user_type != 'teacher':
            return Response({'error': 'Teachers only'}, status=status.HTTP_403_FORBIDDEN)

        teacher_email = str(request.user)
        mode = request.query_params.get('mode', 'months')

        try:
            subject = Subject.objects.select_related('course', 'class_field', 'class_field__department', 'teacher').get(subject_id=pk)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=status.HTTP_404_NOT_FOUND)

        # Verify access: teacher owns the subject or is coordinator
        is_owner = subject.teacher.teacher_email == teacher_email
        is_coordinator = (
            subject.class_field.coordinator and
            subject.class_field.coordinator.teacher_email == teacher_email
        )
        if not is_owner and not is_coordinator:
            return Response({'error': 'Unauthorized'}, status=status.HTTP_403_FORBIDDEN)

        # Get all closed sessions for this subject, ordered by date
        all_sessions = Session.objects.filter(
            period__subject=subject,
            is_active=False,
        ).order_by('date', 'session_id')

        # Get all enrolled students
        enrollments = StudentClass.objects.filter(class_obj=subject.class_field).select_related('student')
        students_list = sorted(enrollments, key=lambda e: e.roll_no)

        # Common metadata
        meta = {
            'subject_name': subject.course.course_name,
            'class_name': str(subject.class_field),
            'department': subject.class_field.department.department_name,
            'batch': subject.class_field.batch,
            'semester': subject.class_field.semester,
            'section': subject.class_field.section,
            'teacher_name': subject.teacher.name,
        }

        if mode == 'months':
            return self._handle_months(all_sessions, meta)
        elif mode == 'monthly':
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            if not month or not year:
                return Response({'error': 'month and year required for monthly mode'}, status=status.HTTP_400_BAD_REQUEST)
            return self._handle_monthly(all_sessions, students_list, meta, int(month), int(year))
        elif mode == 'semester':
            return self._handle_semester(all_sessions, students_list, meta)
        else:
            return Response({'error': 'Invalid mode. Use: months, monthly, or semester'}, status=status.HTTP_400_BAD_REQUEST)

    def _handle_months(self, all_sessions, meta):
        """Return list of months that have session data."""
        months_map = defaultdict(int)
        for session in all_sessions:
            key = (session.date.year, session.date.month)
            months_map[key] += 1

        months = []
        for (year, month), count in sorted(months_map.items()):
            months.append({
                'month': month,
                'year': year,
                'month_name': calendar.month_name[month],
                'total_sessions': count,
            })

        return Response({**meta, 'months': months})

    def _handle_monthly(self, all_sessions, students_list, meta, month, year):
        """Return date-wise attendance register for a specific month."""
        month_sessions = [s for s in all_sessions if s.date.month == month and s.date.year == year]

        if not month_sessions:
            return Response({
                **meta,
                'month': month,
                'year': year,
                'month_name': calendar.month_name[month],
                'sessions': [],
                'students': [],
            })

        session_ids = [s.session_id for s in month_sessions]
        sessions_info = []
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        for s in month_sessions:
            sessions_info.append({
                'date': s.date.isoformat(),
                'day': day_names[s.date.weekday()],
            })

        # Batch-fetch all attendance records for these sessions
        attendance_records = Attendance.objects.filter(
            session_id__in=session_ids
        ).values_list('session_id', 'student_id', 'status')

        # Build lookup: (session_id, student_email) -> status
        att_lookup = {}
        for sid, student_email, att_status in attendance_records:
            att_lookup[(sid, student_email)] = att_status

        students_data = []
        for enrollment in students_list:
            student = enrollment.student
            attendance = []
            total_present = 0
            for s in month_sessions:
                st = att_lookup.get((s.session_id, student.student_email))
                present = 1 if st == 'P' else 0
                attendance.append(present)
                total_present += present

            total_classes = len(month_sessions)
            students_data.append({
                'roll_no': enrollment.roll_no,
                'name': student.name,
                'email': student.student_email,
                'attendance': attendance,
                'total_present': total_present,
                'total_classes': total_classes,
                'percentage': round((total_present / total_classes * 100), 2) if total_classes > 0 else 0,
            })

        return Response({
            **meta,
            'month': month,
            'year': year,
            'month_name': calendar.month_name[month],
            'sessions': sessions_info,
            'students': students_data,
        })

    def _handle_semester(self, all_sessions, students_list, meta):
        """Return month-wise summary for the full semester."""
        # Group sessions by month
        months_sessions = defaultdict(list)
        for s in all_sessions:
            months_sessions[(s.date.year, s.date.month)].append(s)

        sorted_months = sorted(months_sessions.keys())

        if not sorted_months:
            return Response({**meta, 'months': [], 'students': []})

        months_info = []
        for (year, month) in sorted_months:
            months_info.append({
                'month': month,
                'year': year,
                'month_name': calendar.month_name[month],
                'total_sessions': len(months_sessions[(year, month)]),
            })

        # Batch-fetch all attendance records for all sessions
        all_session_ids = [s.session_id for s in all_sessions]
        attendance_records = Attendance.objects.filter(
            session_id__in=all_session_ids
        ).values_list('session_id', 'student_id', 'status')

        # Build lookup: (session_id, student_email) -> status
        att_lookup = {}
        for sid, student_email, att_status in attendance_records:
            att_lookup[(sid, student_email)] = att_status

        # Map session_id -> (year, month)
        session_month_map = {}
        for s in all_sessions:
            session_month_map[s.session_id] = (s.date.year, s.date.month)

        students_data = []
        for enrollment in students_list:
            student = enrollment.student
            monthly_present = defaultdict(int)
            monthly_total = defaultdict(int)

            for s in all_sessions:
                ym = session_month_map[s.session_id]
                st = att_lookup.get((s.session_id, student.student_email))
                monthly_total[ym] += 1
                if st == 'P':
                    monthly_present[ym] += 1

            monthly_stats = []
            total_present = 0
            total_classes = 0
            for ym in sorted_months:
                p = monthly_present[ym]
                t = monthly_total[ym]
                total_present += p
                total_classes += t
                monthly_stats.append({
                    'present': p,
                    'total': t,
                    'percentage': round((p / t * 100), 2) if t > 0 else 0,
                })

            students_data.append({
                'roll_no': enrollment.roll_no,
                'name': student.name,
                'email': student.student_email,
                'monthly_stats': monthly_stats,
                'total_present': total_present,
                'total_classes': total_classes,
                'percentage': round((total_present / total_classes * 100), 2) if total_classes > 0 else 0,
            })

        return Response({
            **meta,
            'months': months_info,
            'students': students_data,
        })